import csv
from pathlib import Path
from openpyxl import Workbook, load_workbook

from utils.config import OUTPUT_DIR

CSV_FILE = OUTPUT_DIR / "DashboardMetrics.csv"
EXCEL_FILE = OUTPUT_DIR / "DashboardMetrics.xlsx"


def initialize_storage(headers):
    """
    Creates CSV and Excel files with headers if they don't exist.
    """

    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # ---------- CSV ----------
        if not CSV_FILE.exists():
            with open(CSV_FILE, "w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(headers)

        # ---------- Excel ----------
        if not EXCEL_FILE.exists():
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "DashboardMetrics"
            sheet.append(headers)
            workbook.save(EXCEL_FILE)

    except Exception as e:
        print(f"Error initializing storage: {e}")
        raise


def save_to_csv(metrics):
    """
    Appends one metrics dictionary to CSV.
    """

    try:
        with open(CSV_FILE, "a", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(file, fieldnames=metrics.keys())
            writer.writerow(metrics)

    except Exception as e:
        print(f"Error saving metrics to CSV: {e}")
        raise


def save_to_excel(metrics):
    """
    Appends one metrics dictionary to Excel.
    """

    try:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

        sheet.append(list(metrics.values()))

        workbook.save(EXCEL_FILE)

    except Exception as e:
        print(f"Error saving metrics to Excel: {e}")
        raise


def save_metrics(metrics):
    """
    Saves metrics to both CSV and Excel.
    """

    try:
        save_to_csv(metrics)
        save_to_excel(metrics)

    except Exception as e:
        print(f"Error saving metrics: {e}")
        raise


def load_csv():
    """
    Returns all stored CSV records.
    """

    try:
        if not CSV_FILE.exists():
            return []

        with open(CSV_FILE, "r", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            return list(reader)

    except Exception as e:
        print(f"Error loading CSV records: {e}")
        return []


def load_excel():
    """
    Returns all stored Excel records.
    """

    try:
        if not EXCEL_FILE.exists():
            return []

        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

        rows = list(sheet.values)

        headers = rows[0]

        data = []

        for row in rows[1:]:
            data.append(dict(zip(headers, row)))

        return data

    except Exception as e:
        print(f"Error loading Excel records: {e}")
        return []