import json
import os
import logging
import re
import time
from pathlib import Path
from typing import List, Optional

from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
# ----------------------------------------------------------
# Logging Configuration
# ----------------------------------------------------------

logger = logging.getLogger(__name__)

if not logger.handlers:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s"
    )

from dotenv import load_dotenv

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(_PROJECT_ROOT / ".env")

# ---------------------------------------------------------------------------
# Gemini Configuration (Jagruthi: gemini-3.5 default, deprecated aliases, 503 fallbacks)
# ---------------------------------------------------------------------------
_DEPRECATED_MODEL_ALIASES = {
    "gemini-2.0-flash": "gemini-3.5-flash",
    "gemini-2.0-flash-lite": "gemini-3.5-flash-lite",
    "gemini-2.5-flash": "gemini-3.5-flash",
    "gemini-2.5-flash-lite": "gemini-3.5-flash-lite",
    "gemini-1.5-flash": "gemini-3.5-flash-lite",
    "gemini-1.5-flash-latest": "gemini-3.5-flash-lite",
    "gemini-1.5-pro": "gemini-3.5-flash",
}

GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-3.5-flash")
GEMINI_MODEL_FALLBACKS = [
    model.strip()
    for model in os.getenv(
        "GEMINI_MODEL_FALLBACKS",
        "gemini-3.5-flash-lite,gemini-3.6-flash",
    ).split(",")
    if model.strip()
]


def _normalize_gemini_model(model_name: str) -> str:
    return _DEPRECATED_MODEL_ALIASES.get(model_name.strip(), model_name.strip())


def _gemini_models_to_try() -> list[str]:
    models: list[str] = []
    for raw_name in [GEMINI_MODEL, *GEMINI_MODEL_FALLBACKS]:
        name = _normalize_gemini_model(raw_name)
        if name and name not in models:
            models.append(name)
    return models


def _parse_retry_delay_seconds(error: Exception) -> float | None:
    candidates = [str(error)]
    response_json = getattr(error, "response_json", None)
    if isinstance(response_json, dict):
        candidates.append(str(response_json))
        details = response_json.get("error", {}).get("details", [])
        for detail in details:
            if isinstance(detail, dict) and "retryDelay" in detail:
                candidates.append(str(detail.get("retryDelay", "")))
    for message in candidates:
        match = re.search(r"retry in ([\d.]+)s", message, re.IGNORECASE)
        if match:
            return float(match.group(1)) + 1.0
        match = re.search(r'"retryDelay":\s*"(\d+)s"', message)
        if match:
            return float(match.group(1)) + 1.0
    return None


def _is_model_not_found(error: Exception) -> bool:
    message = str(error).casefold()
    return "404" in message or "not_found" in message or "is not found" in message


def _should_try_fallback_model(error: Exception) -> bool:
    message = str(error).casefold()
    return (
        _is_model_not_found(error)
        or "503" in message
        or "unavailable" in message
        or "resource_exhausted" in message
        or "429" in message
    )


def _suggested_model_from_error(error: Exception) -> str | None:
    match = re.search(
        r"(?:use models/|update your code to use models/)([a-z0-9._-]+)",
        str(error),
        re.IGNORECASE,
    )
    if match:
        return _normalize_gemini_model(match.group(1))
    return None

def initialize_gemini():

    try:

        logger.info("Loading environment variables")

        load_dotenv(_PROJECT_ROOT / ".env")

        api_key = os.getenv("GEMINI_API_KEY")

        if not api_key:

            logger.error(
                "GEMINI_API_KEY was not found in environment variables"
            )

            raise ValueError(
                "GEMINI_API_KEY was not found in .env"
            )

        logger.info("Gemini API key found")

        client = genai.Client(
            api_key=api_key
        )

        logger.info("Gemini client initialized successfully")

        return client

    except Exception:

        logger.exception(
            "Failed to initialize Gemini client"
        )

        raise


# Initialise only when extraction is requested.  This lets the browser/DOME
# collector and its table comparison complete even if an API key is absent.
client = None

#updated metadata for refreshdate, pagennumber and pagename
class Metadata(BaseModel):
    dashboard_title: Optional[str] = Field(
        None,
        description="Main dashboard title"
    )

    page_name: Optional[str] = Field(
        None,
        description="Currently visible dashboard page or tab name"
    )

    page_number: Optional[str] = Field(
        None,
        description="Visible page number if available"
    )

    data_refresh_date: Optional[str] = Field(
        None,
        description="Dashboard data refresh date or timestamp exactly as displayed"
    )

#filter recognition and selected filter values
class Filter(BaseModel):

    filter_name: str = Field(
        description="Name of the dashboard filter exactly as displayed"
    )

    filter_type: Optional[str] = Field(
        None,
        description=(
            "Type of filter such as Dropdown, List, Radio Button, "
            "Checkbox, Button, or other visible filter type"
        )
    )

    available_values: List[str] = Field(
        default_factory=list,
        description=(
            "All visible values/options available for this filter"
        )
    )

    selected_values: List[str] = Field(
        default_factory=list,
        description=(
            "Currently selected filter values. "
            "Return all selected values for multi-select filters"
        )
    )


class KPICard(BaseModel):

    name: str = Field(
        description="KPI name or label exactly as displayed"
    )

    value: Optional[str] = Field(
        None,
        description="Current KPI value exactly as displayed"
    )

    previous_value: Optional[str] = Field(
        None,
        description="Previous or comparison KPI value if visibly displayed"
    )

    variance: Optional[str] = Field(
        None,
        description="Variance, percentage change, or comparison value if displayed"
    )

class ChartDataPoint(BaseModel):

    category: Optional[str] = Field(
        None,
        description="Category or label represented by this data point"
    )

    value: Optional[str] = Field(
        None,
        description="Value represented by this data point"
    )

    series: Optional[str] = Field(
        None,
        description="Series name if the chart contains multiple series"
    )

    block_color: Optional[str] = Field(
        None,
        description="Visible color of the chart or treemap block"
    )

    block_position: Optional[str] = Field(
        None,
        description="Approximate spatial position of the data block in the visual"
    )

class Chart(BaseModel):

    visual_id: Optional[str] = Field(
        None,
        description="Identifier for this visual within the dashboard"
    )

    chart_title: Optional[str] = Field(
        None,
        description="Chart title exactly as displayed"
    )

    chart_type: Optional[str] = Field(
        None,
        description="Type of visual such as bar, line, pie, table, etc."
    )

    data: List[ChartDataPoint] = Field(
        default_factory=list,
        description="Visible chart data points"
    )

    series: List[str] = Field(
        default_factory=list,
        description="Visible data series or measures"
    )

    block_color: Optional[str] = Field(
        None,
        description="Dominant or clearly identifiable visual color"
    )

    block_position: Optional[str] = Field(
        None,
        description="Approximate position of the visual on the dashboard"
    )


class Table(BaseModel):

    table_title: Optional[str] = Field(
        None,
        description="Table title if visible"
    )

    columns: List[str] = Field(
        default_factory=list,
        description="Visible table column names"
    )

    rows: List[List[str]] = Field(
        default_factory=list,
        description="Visible table rows"
    )

class DashboardExtraction(BaseModel):

    metadata: Metadata = Field(
        default_factory=Metadata
    )

    filters: List[Filter] = Field(
        default_factory=list
    )

    kpi_cards: List[KPICard] = Field(
        default_factory=list
    )

    charts: List[Chart] = Field(
        default_factory=list
    )

    tables: List[Table] = Field(
        default_factory=list
    )


# ---------------------------------------------------------------------------
# Extraction Function
# ---------------------------------------------------------------------------
def extract_dashboard_json(image_path: Path) -> dict:
    """Extract structured dashboard data from image bytes via Gemini, returning a safe fallback on JSON truncation."""
    global client
    logger.info("Starting dashboard extraction: %s", image_path.name)

    if not image_path.exists():
        logger.error("Dashboard image not found: %s", image_path)
        raise FileNotFoundError(f"Dashboard image not found: {image_path}")

    image_bytes = image_path.read_bytes()

    prompt = """
        ==========================================================
        DASHBOARD INFORMATION EXTRACTION
        ==========================================================

        Analyze the entire dashboard image and extract all visible
        dashboard information that can be reliably identified.

        Do NOT invent values that are not visible.

        ==========================================================
        FILTER DETECTION
        ==========================================================

        Identify EVERY visible filter on the dashboard.

        A filter may appear as:

        - Dropdown
        - Slicer
        - List selector
        - Radio button group
        - Button-style selector
        - Other selectable filter control

        IMPORTANT:

        The word "slicer" may appear in the underlying implementation,
        but in the extracted result ALWAYS call it a "filter".

        For every filter, extract:

        1. Filter name
        2. Filter type
        3. All currently visible/available filter values
        4. Currently selected filter value(s)

        Examples:

        Dropdown:

        Filter name:
        Relative Time

        Filter type:
        Dropdown

        Available values:
        - Current Period
        - Previous Period
        - Previous 3 Periods
        - Previous 6 Periods

        Selected values:
        - Previous 6 Periods


        Radio button:

        Filter name:
        Metrics

        Filter type:
        Radio Button

        Available values:
        - Associate Count
        - Hours Spent
        - Average Completion Rate

        Selected values:
        - Average Completion Rate


        ==========================================================
        FILTER RULES
        ==========================================================

        1. Detect every visible filter independently.

        2. Do not assume all dashboards have the same filters.

        3. Do not hardcode filter names.

        4. Do not assume filter values.

        5. Read the actual displayed values from the image.

        6. If multiple values are selected, return all selected values.

        7. If "All" is selected, return:
        selected_values = ["All"]

        8. If the dashboard displays "Multiple selections", use the
        actual visible selected values when they can be identified.

        9. If the filter is a radio-button group, identify it as:
        "Radio Button"

        10. If the filter is a dropdown, identify it as:
            "Dropdown"

        11. If the filter is a list selector, identify it as:
            "List"

        12. If it is a button-style filter, identify it as:
            "Buttons"

        13. Count the values that are actually visible.

        14. Never treat ordinary dashboard buttons, navigation buttons,
        bookmarks, images, or links as filters unless they clearly
        function as a selectable filter.

        15. Do not treat page navigation controls as filters.

        16. Do not treat bookmark controls as filters.

        ==========================================================
        REFRESH DATE
        ==========================================================

        Find the dashboard's data refresh date or refresh timestamp.

        Look for text such as:

        - Refresh Date
        - Data Refresh Date
        - Last Refresh
        - Last Refreshed
        - Data Updated
        - Updated
        - Refresh Time

        Extract the exact date/time as displayed.

        Do not infer or calculate the refresh date.

        If no refresh date is visible, return null.

        ==========================================================
        PAGE INFORMATION
        ==========================================================

        Identify:

        - Dashboard title
        - Current page/tab name
        - Page number if visibly available

        Do not infer page numbers.

        ==========================================================
        KPI EXTRACTION
        ==========================================================

        Identify EVERY visible KPI card before extracting any charts. A KPI
        card is a compact label/value visual such as "Total Sales 1,250" or a
        card with a prior value/variance; it is not an axis label, legend,
        tooltip, table cell, chart title, navigation control, or slicer value.

        Work left-to-right and top-to-bottom. First count the KPI cards and
        then return one kpi_cards object per card. Keep the label and value
        paired from the same card. If a value cannot be read sharply, leave it
        null rather than borrowing a nearby chart value. Preserve display
        punctuation, currency symbols, percentage symbols, and suffixes.

        Identify visible KPI cards and extract:

        - KPI name
        - Current value
        - Prior period value if visible
        - Variance if visible

        Do not invent missing values.

        ==========================================================
        VISUAL EXTRACTION
        ==========================================================

        Identify visible charts and tables and extract their
        relevant information.

        ==========================================================
        TREEMAP EXTRACTION (COLOR-BASED)
        ==========================================================

        Treat every distinct colored region as one independent
        treemap data object.
        """

    response = None
    models = _gemini_models_to_try()

    for model_name in models:
        try:
            if client is None:
                client = initialize_gemini()
            response = client.models.generate_content(
                model=model_name,
                contents=[
                    types.Part.from_bytes(data=image_bytes, mime_type="image/png"),
                    prompt,
                ],
                config=types.GenerateContentConfig(
                    temperature=0,
                    response_mime_type="application/json",
                    response_schema=DashboardExtraction,
                ),
            )
            if response and response.text:
                break
        except Exception as e:
            logger.warning("Gemini model %s failed: %s", model_name, e)

    if not response or not response.text:
        logger.error("Gemini returned empty or invalid response across all models.")
        return {"metadata": {}, "filters": [], "kpi_cards": [], "charts": [], "tables": []}

    raw_text = response.text.strip()
    cleaned_text = re.sub(r"^```(?:json)?\s*", "", raw_text, flags=re.IGNORECASE)
    cleaned_text = re.sub(r"\s*```$", "", cleaned_text)

    try:
        extracted_data = json.loads(cleaned_text)
        logger.info("Gemini response successfully converted to JSON")
        return extracted_data
    except json.JSONDecodeError as e:
        logger.error("Gemini output was truncated or invalid JSON for %s: %s", image_path.name, e)
        return {
            "metadata": {},
            "filters": [],
            "kpi_cards": [],
            "charts": [],
            "tables": [],
            "error": f"Truncated AI JSON output ({len(raw_text)} chars)"
        }

#dashboard Summary
def build_dashboard_summary(extracted_data: dict) -> dict:
    """
    Build calculated dashboard counts and summary information.

    Counts are calculated in Python rather than asking Gemini
    to calculate them.
    """

    logger.info(
        "Building dashboard summary"
    )

    try:

        metadata = extracted_data.get(
            "metadata",
            {}
        )

        filters = extracted_data.get(
            "filters",
            []
        )

        kpis = extracted_data.get(
            "kpi_cards",
            []
        )

        charts = extracted_data.get(
            "charts",
            []
        )

        tables = extracted_data.get(
            "tables",
            []
        )

        filter_count = len(filters)
        kpi_count = len(kpis)
        chart_count = len(charts)
        table_count = len(tables)

        visual_count = (
            kpi_count
            + chart_count
            + table_count
        )

        summary = {

            "filter_count": filter_count,

            "kpi_count": kpi_count,

            "chart_count": chart_count,

            "table_count": table_count,

            "visual_count": visual_count,

            "page_name": metadata.get(
                "page_name"
            ),

            "page_number": metadata.get(
                "page_number"
            ),

            "refresh_date": metadata.get(
                "data_refresh_date"
            ),

            "dashboard_title": metadata.get(
                "dashboard_title"
            ),
        }

        logger.info(
            "Dashboard summary created | "
            "filters=%d | KPIs=%d | charts=%d | tables=%d | visuals=%d",
            filter_count,
            kpi_count,
            chart_count,
            table_count,
            visual_count,
        )

        return summary

    except Exception:

        logger.exception(
            "Failed to build dashboard summary"
        )

        raise
#Filter Comparison
# Jagruthi: case-insensitive filter name matching for source vs target dashboards.
def compare_filters(source_data: dict, target_data: dict) -> list:
    """Compare filters between source and target dashboards using safe type normalization."""
    logger.info("Starting source-vs-target filter comparison")

    def _normalize_str(val) -> str:
        return " ".join(str(val or "").casefold().split())

    def _to_normalized_set(val) -> set:
        if isinstance(val, list):
            return {_normalize_str(item) for item in val if str(item).strip()}
        elif isinstance(val, str) and val.strip():
            return {_normalize_str(val)}
        return set()

    source_filters = source_data.get("filters", []) or []
    target_filters = target_data.get("filters", []) or []

    source_map = {
        _normalize_str(f.get("filter_name")): f
        for f in source_filters if isinstance(f, dict) and f.get("filter_name")
    }
    target_map = {
        _normalize_str(f.get("filter_name")): f
        for f in target_filters if isinstance(f, dict) and f.get("filter_name")
    }

    results = []
    all_filter_names = sorted(set(source_map.keys()) | set(target_map.keys()))

    for name in all_filter_names:
        source = source_map.get(name)
        target = target_map.get(name)
        display_name = (source or {}).get("filter_name") or (target or {}).get("filter_name") or name

        if source is None:
            results.append({
                "filter_name": display_name,
                "source_type": None,
                "target_type": target.get("filter_type"),
                "source_selected": [],
                "target_selected": target.get("selected_values", []) or [],
                "status": "Missing in Source"
            })
            continue

        if target is None:
            results.append({
                "filter_name": display_name,
                "source_type": source.get("filter_type"),
                "target_type": None,
                "source_selected": source.get("selected_values", []) or [],
                "target_selected": [],
                "status": "Missing in Target"
            })
            continue

        src_sel_set = _to_normalized_set(source.get("selected_values"))
        tgt_sel_set = _to_normalized_set(target.get("selected_values"))
        src_val_set = _to_normalized_set(source.get("available_values"))
        tgt_val_set = _to_normalized_set(target.get("available_values"))

        if src_sel_set != tgt_sel_set:
            status = "Selection Changed"
        elif src_val_set != tgt_val_set:
            status = "Available Values Changed"
        elif _normalize_str(source.get("filter_type")) != _normalize_str(target.get("filter_type")):
            status = "Type Changed"
        else:
            status = "Match"

        results.append({
            "filter_name": display_name,
            "source_type": source.get("filter_type"),
            "target_type": target.get("filter_type"),
            "source_selected": list(src_sel_set),
            "target_selected": list(tgt_sel_set),
            "status": status
        })

    return results

#visual comparison
def compare_visuals(
    source_data: dict,
    target_data: dict
) -> list:
    """
    Compare visuals between source and target dashboards.

    Checks:
    - Missing visuals
    - Chart title
    - Chart type
    - Chart data
    """

    logger.info(
        "Starting visual comparison"
    )

    try:

        # --------------------------------------------------
        # Extract charts
        # --------------------------------------------------

        source_charts = source_data.get(
            "charts",
            []
        )

        target_charts = target_data.get(
            "charts",
            []
        )

        logger.info(
            "Visual counts | source=%d | target=%d",
            len(source_charts),
            len(target_charts)
        )

        # --------------------------------------------------
        # Create lookup maps
        # --------------------------------------------------

        source_map = {
            chart.get("visual_id"): chart
            for chart in source_charts
            if chart.get("visual_id")
        }

        target_map = {
            chart.get("visual_id"): chart
            for chart in target_charts
            if chart.get("visual_id")
        }

        logger.info(
            "Visual lookup maps created | "
            "source=%d | target=%d",
            len(source_map),
            len(target_map)
        )

        results = []

        all_visual_ids = sorted(
            set(source_map.keys())
            |
            set(target_map.keys())
        )

        logger.info(
            "Total unique visuals to compare: %d",
            len(all_visual_ids)
        )

        # --------------------------------------------------
        # Compare visuals
        # --------------------------------------------------

        for visual_id in all_visual_ids:

            source = source_map.get(
                visual_id
            )

            target = target_map.get(
                visual_id
            )

            # ----------------------------------------------
            # Missing in Source
            # ----------------------------------------------

            if source is None:

                logger.warning(
                    "Visual '%s' missing in source",
                    visual_id
                )

                results.append({
                    "visual_id": visual_id,
                    "status": "Missing in Source"
                })

                continue

            # ----------------------------------------------
            # Missing in Target
            # ----------------------------------------------

            if target is None:

                logger.warning(
                    "Visual '%s' missing in target",
                    visual_id
                )

                results.append({
                    "visual_id": visual_id,
                    "status": "Missing in Target"
                })

                continue

            # ----------------------------------------------
            # Compare properties
            # ----------------------------------------------

            differences = []

            if (
                source.get("chart_title")
                != target.get("chart_title")
            ):
                differences.append("Title")

            if (
                source.get("chart_type")
                != target.get("chart_type")
            ):
                differences.append("Chart Type")

            if (
                source.get("data")
                != target.get("data")
            ):
                differences.append("Data")

            status = (
                "Match"
                if not differences
                else "Changed: " + ", ".join(differences)
            )

            logger.info(
                "Visual comparison | visual_id=%s | status=%s",
                visual_id,
                status
            )

            results.append({
                "visual_id": visual_id,
                "source_title": source.get("chart_title"),
                "target_title": target.get("chart_title"),
                "source_type": source.get("chart_type"),
                "target_type": target.get("chart_type"),
                "status": status
            })

        # --------------------------------------------------
        # Final summary
        # --------------------------------------------------

        match_count = sum(
            1
            for result in results
            if result.get("status") == "Match"
        )

        logger.info(
            "Visual comparison completed | "
            "total=%d | matches=%d",
            len(results),
            match_count
        )

        return results

    except Exception:

        logger.exception(
            "Visual comparison failed"
        )

        raise
# ---------------------------------------------------------------------------
# Excel Formatter & Exporter
# ---------------------------------------------------------------------------
def export_json_to_excel(extracted_data: dict, output_file_path: Path):
    """Format and export extracted JSON structure to styled Excel workbook."""
    wb = openpyxl.Workbook()
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell_font = Font(name=font_family, size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def format_sheet_header(ws, headers):
        ws.views.sheetView[0].showGridLines = True
        ws.row_dimensions[1].height = 26
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def auto_fit_columns(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # 1. Filters / Slicers Worksheet (Fixes key mismatch)
    filters = extracted_data.get("filters", []) or extracted_data.get("slicers", [])
    if filters:
        ws = wb.create_sheet(title="Slicers")
        format_sheet_header(ws, ["Filter Name", "Filter Type", "Selected Value(s)", "Available Values"])
        row = 2
        for f in filters:
            sel = f.get("selected_values") or f.get("selected_value") or "All / Unselected"
            if isinstance(sel, list):
                sel = ", ".join(map(str, sel))
            avail = f.get("available_values", [])
            if isinstance(avail, list):
                avail = ", ".join(map(str, avail))

            vals = [f.get("filter_name"), f.get("filter_type") or "N/A", str(sel), str(avail)]
            for col_idx, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = cell_font
                c.border = thin_border
            ws.row_dimensions[row].height = 20
            row += 1
        auto_fit_columns(ws)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])

    wb.save(output_file_path)

# ---------------------------------------------------------------------------
# Process Folder
# ---------------------------------------------------------------------------
def process_folder(
    input_folder: str,
    output_folder: str
):
    """
    Process all dashboard images in a folder.

    For each dashboard:
    1. Extract information using Gemini
    2. Save JSON
    3. Generate Excel
    """

    logger.info(
        "Starting dashboard folder processing"
    )

    try:

        # --------------------------------------------------
        # Prepare paths
        # --------------------------------------------------

        input_path = Path(
            input_folder
        )

        output_path = Path(
            output_folder
        )

        logger.info(
            "Input folder: %s",
            input_path
        )

        logger.info(
            "Output folder: %s",
            output_path
        )

        if not input_path.exists():

            logger.error(
                "Input folder does not exist: %s",
                input_path
            )

            raise FileNotFoundError(
                f"Input folder does not exist: {input_path}"
            )

        output_path.mkdir(
            exist_ok=True
        )

        # --------------------------------------------------
        # Find dashboard images
        # --------------------------------------------------

        image_extensions = {
            ".png",
            ".jpg",
            ".jpeg"
        }

        images = sorted(
            [
                f
                for f in input_path.iterdir()
                if (
                    f.is_file()
                    and f.suffix.lower()
                    in image_extensions
                )
            ]
        )

        logger.info(
            "Found %d dashboard image(s)",
            len(images)
        )

        print(
            f"Found {len(images)} images to process\n"
        )

        # --------------------------------------------------
        # Process each dashboard
        # --------------------------------------------------

        for image in images:

            logger.info(
                "Processing dashboard image: %s",
                image.name
            )

            try:

                # ------------------------------------------
                # 1. Gemini extraction
                # ------------------------------------------

                logger.info(
                    "Starting Gemini extraction: %s",
                    image.name
                )

                extracted_json = extract_dashboard_json(
                    image
                )

                logger.info(
                    "Gemini extraction completed: %s",
                    image.name
                )

                # ------------------------------------------
                # 2. Save JSON
                # ------------------------------------------

                json_output_file = (
                    output_path
                    / f"{image.stem}.json"
                )

                with open(
                    json_output_file,
                    "w",
                    encoding="utf-8"
                ) as f:

                    json.dump(
                        extracted_json,
                        f,
                        indent=2
                    )

                logger.info(
                    "JSON saved: %s",
                    json_output_file
                )

                print(
                    f"Saved JSON: "
                    f"{json_output_file.name}"
                )

                # ------------------------------------------
                # 3. Export Excel
                # ------------------------------------------

                excel_output_file = (
                    output_path
                    / f"{image.stem}.xlsx"
                )

                logger.info(
                    "Starting Excel export: %s",
                    image.name
                )

                export_json_to_excel(
                    extracted_json,
                    excel_output_file
                )

                logger.info(
                    "Excel saved: %s",
                    excel_output_file
                )

                print(
                    f"Saved Excel: "
                    f"{excel_output_file.name}\n"
                )

            except Exception:

                logger.exception(
                    "Failed processing dashboard: %s",
                    image.name
                )

                print(
                    f"Failed processing "
                    f"{image.name}\n"
                )

                # Continue with next dashboard
                continue

        logger.info(
            "Dashboard folder processing completed"
        )

    except Exception:

        logger.exception(
            "Dashboard folder processing failed"
        )

        raise


if __name__ == "__main__":
    process_folder(
        input_folder="dashboard",
        output_folder="extracted_results",
    )
    print("Done. All JSON and Excel reports generated.")
