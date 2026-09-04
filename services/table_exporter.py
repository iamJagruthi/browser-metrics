"""
table_exporter.py

Service responsible only for exporting Power BI
table and matrix visuals.

This module:
- does not launch its own browser
- does not navigate dashboards
- does not compare source and target
- does not use AI/Gemini
"""

"""
Table exporter service.

Responsibilities:
- Receive a detected Table/Matrix visual.
- Use Power BI native Export Data functionality.
- Parse downloaded CSV/XLSX files.
- Return structured table data.

This file:
- Does NOT use Gemini/LLM.
- Does NOT compare Source vs Target.
- Does NOT run as a standalone application.
- Does NOT launch browsers itself.
"""


import csv
import json
import logging
import re
import uuid

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from utils.config import OUTPUT_DIR


logger = logging.getLogger(__name__)


# ---------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------

VISUAL_SELECTOR = (
    ".visualContainer, "
    "[data-visual-container]"
)

MAX_EXPORT_RETRIES = 3

MENU_TIMEOUT = 8_000

DOWNLOAD_TIMEOUT = 20_000


# ---------------------------------------------------------
# OUTPUT DIRECTORIES
# ---------------------------------------------------------

EXPORT_DIR = OUTPUT_DIR / "table_exports"
RAW_DIR = EXPORT_DIR / "raw"
JSON_DIR = EXPORT_DIR / "json"

class TableExporter:

    def __init__(
        self,
        page,
        output_dir: str | Path = "output/table_exports",
    ):
        self.page = page
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

    async def export_table_visual(
        self,
        locator,
        visual_metadata: dict[str, Any],
        dashboard_name: str,
    ) -> dict[str, Any]:
        """
        Export one already-identified Power BI
        table or matrix visual.

        The caller is responsible for detecting
        that this is a table/matrix.
        """

        title = (
            visual_metadata.get("title")
            or f"table_visual_{visual_metadata.get('index', 'unknown')}"
        )

        logger.info(
            "Starting table export | dashboard=%s | title=%s",
            dashboard_name,
            title,
        )

        result = {
            "title": title,
            "visual_id": visual_metadata.get("id"),
            "index": visual_metadata.get("index"),
            "is_table": visual_metadata.get("is_table", False),
            "is_matrix": visual_metadata.get("is_matrix", False),
            "status": "not_exported",
            "columns": [],
            "rows": [],
            "row_count": 0,
            "file_path": None,
            "error": None,
        }

        try:
            export_path = await self._export_visual_data(
                locator=locator,
                title=title,
                dashboard_name=dashboard_name,
            )

            if not export_path:
                result["status"] = "export_failed"
                result["error"] = (
                    "Power BI export did not produce a file."
                )

                logger.warning(
                    "Table export produced no file | title=%s",
                    title,
                )

                return result

            result["file_path"] = str(export_path)

            parsed_data = await self._parse_export_file(
                export_path
            )

            result["columns"] = parsed_data.get(
                "columns",
                [],
            )

            result["rows"] = parsed_data.get(
                "rows",
                [],
            )

            result["row_count"] = len(
                result["rows"]
            )

            result["status"] = "success"

            logger.info(
                "Table export completed | title=%s | rows=%d | columns=%d",
                title,
                result["row_count"],
                len(result["columns"]),
            )

            return result

        except Exception as exc:
            logger.exception(
                "Table export failed | title=%s",
                title,
            )

            result["status"] = "failed"
            result["error"] = str(exc)

            return result

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_filename(value: str, fallback: str) -> str:
    value = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        value,
    ).strip("_.")

    return value or fallback


def _clean(value: Any) -> str:
    return " ".join(
        str(value if value is not None else "").split()
    )


# ---------------------------------------------------------------------------
# Export parsing
# ---------------------------------------------------------------------------

def _read_export(path: Path) -> dict[str, Any]:

    if path.suffix.lower() == ".csv":

        with path.open(
            "r",
            encoding="utf-8-sig",
            newline="",
        ) as handle:
            values = list(csv.reader(handle))

    elif path.suffix.lower() in {".xlsx", ".xlsm"}:

        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
        )

        sheet = workbook.active

        values = [
            list(row)
            for row in sheet.iter_rows(
                values_only=True
            )
        ]

        workbook.close()

    else:
        raise ValueError(
            f"Unsupported export format: {path.suffix}"
        )

    values = [
        [
            _clean(value)
            for value in row
        ]
        for row in values
        if any(
            _clean(value)
            for value in row
        )
    ]

    if not values:
        return {
            "columns": [],
            "rows": [],
            "row_count": 0,
        }

    width = max(
        len(row)
        for row in values
    )

    values = [
        row + [""] * (width - len(row))
        for row in values
    ]

    return {
        "columns": values[0],
        "rows": values[1:],
        "row_count": max(
            0,
            len(values) - 1,
        ),
    }


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------

def _normalise_table(
    table: dict[str, Any],
) -> tuple[tuple[str, ...], ...]:

    columns = tuple(
        _clean(x).casefold()
        for x in table.get(
            "columns",
            [],
        )
    )

    rows = tuple(
        tuple(
            _clean(value).casefold()
            for value in row
        )
        for row in table.get(
            "rows",
            [],
        )
    )

    return (columns,) + rows


def compare_tables(
    source: dict[str, Any],
    target: dict[str, Any],
) -> dict[str, Any]:

    source_columns = source.get(
        "columns",
        [],
    )

    target_columns = target.get(
        "columns",
        [],
    )

    source_rows = source.get(
        "rows",
        [],
    )

    target_rows = target.get(
        "rows",
        [],
    )

    result = {
        "status": "match",
        "columns_match": (
            source_columns == target_columns
        ),
        "row_count_source": len(source_rows),
        "row_count_target": len(target_rows),
        "row_count_match": (
            len(source_rows)
            == len(target_rows)
        ),
        "exact_data_match": (
            _normalise_table(source)
            == _normalise_table(target)
        ),
        "differences": [],
    }

    if not result["columns_match"]:

        result["status"] = "mismatch"

        result["differences"].append({
            "type": "columns",
            "source": source_columns,
            "target": target_columns,
        })

    if not result["row_count_match"]:

        result["status"] = "mismatch"

        result["differences"].append({
            "type": "row_count",
            "source": len(source_rows),
            "target": len(target_rows),
        })

    if not result["exact_data_match"]:

        result["status"] = "mismatch"

        max_rows = max(
            len(source_rows),
            len(target_rows),
        )

        for row_index in range(max_rows):

            left = (
                source_rows[row_index]
                if row_index < len(source_rows)
                else []
            )

            right = (
                target_rows[row_index]
                if row_index < len(target_rows)
                else []
            )

            max_cols = max(
                len(left),
                len(right),
            )

            for col_index in range(max_cols):

                source_value = (
                    left[col_index]
                    if col_index < len(left)
                    else ""
                )

                target_value = (
                    right[col_index]
                    if col_index < len(right)
                    else ""
                )

                if (
                    _clean(source_value).casefold()
                    !=
                    _clean(target_value).casefold()
                ):

                    result["differences"].append({
                        "type": "cell",
                        "row": row_index + 1,
                        "column": col_index + 1,
                        "source": source_value,
                        "target": target_value,
                    })

                    if len(
                        result["differences"]
                    ) >= 100:

                        result["differences"].append({
                            "type": "truncated",
                            "message": (
                                "Only the first "
                                "100 differences are shown."
                            ),
                        })

                        return result

    return result


# ---------------------------------------------------------------------------
# Visual detection
# ---------------------------------------------------------------------------
# NOTE:
# This function is retained for debugging/backward compatibility.
#
# Normal project flow does NOT call this function.
#
# Table/Matrix detection is performed by:
# services/visual_data_exporter.py
#
# Detected table metadata is then passed to:
# export_table_visuals()

async def _find_table_visuals(
    page,
) -> list[dict[str, Any]]:
    """
    Detect ONLY Power BI Table/Matrix visuals.

    We deliberately do NOT use broad DOM heuristics such as:
        [class*="matrix"]
        [role="row"]
        [role="grid"]

    Those caused Cards/Images/other visuals to be
    incorrectly classified as tabular.
    """

    visuals = page.locator(
        VISUAL_SELECTOR
    )

    count = await visuals.count()

    tables: list[dict[str, Any]] = []

    for index in range(count):

        locator = visuals.nth(index)

        try:
            if not await locator.is_visible():
                continue

            info = await locator.evaluate(
                """
                node => {

                    const clean = value =>
                        (value || "")
                            .replace(/\\s+/g, " ")
                            .trim();

                    const ariaRole =
                        clean(
                            node.getAttribute(
                                "aria-roledescription"
                            )
                        );

                    const ariaLabel =
                        clean(
                            node.getAttribute(
                                "aria-label"
                            )
                        );

                    const dataVisualType =
                        clean(
                            node.getAttribute(
                                "data-visual-type"
                            )
                        );

                    const titleNode =
                        node.querySelector(
                            '.visualTitle, ' +
                            '[class*="visualTitle" i], ' +
                            '[data-visual-title]'
                        );

                    const title =
                        clean(
                            titleNode?.innerText
                        ) ||
                        ariaLabel;

                    const isTable =
                        /^table$/i.test(
                            ariaRole
                        ) ||
                        /^table$/i.test(
                            dataVisualType
                        );

                    const isMatrix =
                        /^matrix$/i.test(
                            ariaRole
                        ) ||
                        /^matrix$/i.test(
                            dataVisualType
                        );

                    return {
                        title,
                        aria_label: ariaLabel,
                        aria_role: ariaRole,
                        data_visual_type: dataVisualType,
                        is_table: isTable,
                        is_matrix: isMatrix,
                        is_tabular: (
                            isTable ||
                            isMatrix
                        )
                    };
                }
                """
            )

        except Exception as exc:

            logger.debug(
                "Could not inspect visual #%s: %s",
                index,
                exc,
            )

            continue

        if not info["is_tabular"]:
            continue

        tables.append({
            "index": index,
            "title": (
                info["title"]
                or f"Table {index + 1}"
            ),
            "visual_type": (
                info["aria_role"]
                or info["data_visual_type"]
                or "unknown"
            ),
            "aria_label": info["aria_label"],
        })

    return tables


# ---------------------------------------------------------------------------
# Visual lookup
# ---------------------------------------------------------------------------

async def _get_visual_locator(
    page,
    visual: dict[str, Any],
):
    """
    Re-find the visual every time.

    Power BI frequently rebuilds visual DOM nodes,
    so retaining an old locator can result in:

        element was detached from the DOM
    """

    visuals = page.locator(
        VISUAL_SELECTOR
    )

    aria_label = _clean(
        visual.get("aria_label")
    )

    visual_type = _clean(
        visual.get("visual_type")
    )

    # First try to locate using aria-label + role.
    if aria_label:

        for index in range(
            await visuals.count()
        ):

            candidate = visuals.nth(index)

            try:

                if not await candidate.is_visible():
                    continue

                candidate_info = await candidate.evaluate(
                    """
                    node => ({
                        ariaLabel:
                            (
                                node.getAttribute(
                                    "aria-label"
                                ) || ""
                            ).trim(),

                        ariaRole:
                            (
                                node.getAttribute(
                                    "aria-roledescription"
                                ) || ""
                            ).trim()
                    })
                    """
                )

                if (
                    _clean(
                        candidate_info["ariaLabel"]
                    )
                    == aria_label
                    and (
                        not visual_type
                        or _clean(
                            candidate_info["ariaRole"]
                        ).casefold()
                        == visual_type.casefold()
                    )
                ):

                    return candidate

            except Exception:
                continue

    # Fallback to original index.
    original_index = visual.get(
        "index"
    )

    if original_index is not None:

        if (
            original_index
            < await visuals.count()
        ):
            return visuals.nth(
                original_index
            )

    return None


# ---------------------------------------------------------------------------
# Overlay handling
# ---------------------------------------------------------------------------

async def _close_open_overlays(page) -> None:
    """
    Close stale Power BI menus/dialogs before
    interacting with the next visual.
    """

    try:
        await page.keyboard.press(
            "Escape"
        )

        await page.wait_for_timeout(
            250
        )

        # A second Escape helps when a nested
        # Power BI/Angular overlay is present.
        await page.keyboard.press(
            "Escape"
        )

        await page.wait_for_timeout(
            250
        )

    except Exception as exc:

        logger.debug(
            "Could not close overlay: %s",
            exc,
        )


# ---------------------------------------------------------------------------
# More options
# ---------------------------------------------------------------------------

async def _open_more_options(
    page,
    visual: dict[str, Any],
) -> bool:
    """
    Open the Power BI More options button.

    Uses the stable Power BI selector:

        data-testid="visual-more-options-btn"
    """

    for attempt in range(
        1,
        MAX_EXPORT_RETRIES + 1,
    ):

        try:

            await _close_open_overlays(
                page
            )

            locator = await _get_visual_locator(
                page,
                visual,
            )

            if locator is None:

                logger.warning(
                    "Visual could not be found | "
                    "title=%s",
                    visual["title"],
                )

                continue

            # Make sure the visual is in view.
            try:
                await locator.scroll_into_view_if_needed(
                    timeout=3_000
                )
            except Exception:
                pass

            await page.wait_for_timeout(
                300
            )

            # Reacquire after scrolling because
            # Power BI can rebuild the DOM.
            locator = await _get_visual_locator(
                page,
                visual,
            )

            if locator is None:
                continue

            # Hover is needed because Power BI
            # often reveals the visual toolbar only
            # after the visual is active.
            try:

                await locator.hover(
                    timeout=3_000
                )

            except Exception as hover_exc:

                logger.debug(
                    "Normal hover failed "
                    "(attempt %s): %s",
                    attempt,
                    hover_exc,
                )

                # The visual may have an SVG/path
                # intercepting the pointer. Force hover
                # bypasses Playwright's hit-target check.
                try:

                    await locator.hover(
                        timeout=3_000,
                        force=True,
                    )

                except Exception as force_hover_exc:

                    logger.debug(
                        "Force hover failed: %s",
                        force_hover_exc,
                    )

            await page.wait_for_timeout(
                500
            )

            # Reacquire again after hover.
            locator = await _get_visual_locator(
                page,
                visual,
            )

            if locator is None:
                continue

            menu = locator.locator(
                "button[data-testid='visual-more-options-btn'], "
                "button[aria-label='More options'], "
                "[role='button'][aria-label='More options']"
            ).first

            if await menu.count() == 0:

                logger.warning(
                    "More options button not found | "
                    "title=%s | attempt=%s",
                    visual["title"],
                    attempt,
                )

                continue

            try:

                await menu.click(
                    timeout=3_000
                )

            except Exception as click_exc:

                logger.debug(
                    "Normal More options click "
                    "failed: %s",
                    click_exc,
                )

                # Power BI sometimes has an SVG/path
                # intercepting the pointer. Force click
                # is a safer fallback than changing the
                # page DOM.
                try:

                    await menu.click(
                        timeout=3_000,
                        force=True,
                    )

                except Exception as force_click_exc:

                    logger.debug(
                        "Force More options click "
                        "failed: %s",
                        force_click_exc,
                    )

                    # Last fallback: invoke the DOM click.
                    await menu.evaluate(
                        """
                        element => element.click()
                        """
                    )

            # Wait for the menu to actually render.
            try:

                await page.get_by_role(
                    "menu"
                ).wait_for(
                    state="visible",
                    timeout=MENU_TIMEOUT,
                )

            except Exception:

                # Some Power BI versions don't expose
                # the menu with role="menu".
                await page.wait_for_timeout(
                    500
                )

            return True

        except Exception as exc:

            logger.debug(
                "More options attempt %s failed "
                "for '%s': %s",
                attempt,
                visual["title"],
                exc,
            )

            await _close_open_overlays(
                page
            )

            await page.wait_for_timeout(
                500
            )

    return False


# ---------------------------------------------------------------------------
# Export Data menu item
# ---------------------------------------------------------------------------

async def _find_export_data_item(
    page,
):
    """
    Locate Power BI's Export data menu item.

    Try semantic role first, then text fallback.
    """

    # Preferred.
    try:

        item = page.get_by_role(
            "menuitem",
            name=re.compile(
                r"^\s*export data\s*$",
                re.I,
            ),
        ).first

        if await item.count() > 0:
            return item

    except Exception:
        pass

    # Exact text fallback.
    try:

        item = page.get_by_text(
            re.compile(
                r"^\s*export data\s*$",
                re.I,
            )
        ).first

        if await item.count() > 0:
            return item

    except Exception:
        pass

    # Broader fallback.
    try:

        item = page.get_by_text(
            re.compile(
                r"export\s+data",
                re.I,
            )
        ).first

        if await item.count() > 0:
            return item

    except Exception:
        pass

    return None
#responsible for clicking the export buttona nd choosing full, summarized, underlying data
async def _handle_export_dialog(page) -> dict[str, Any]:
    """
    Handle Power BI's export-options dialog.

    Priority:
        1. Data with current layout
        2. Summarized data

    Returns information about which export mode was actually used.
    """

    dialog = None

    # ---------------------------------------------------------
    # Find Power BI export dialog
    # ---------------------------------------------------------
    try:
        candidate = page.get_by_role(
            "dialog"
        ).filter(
            has_text=re.compile(
                r"which data do you want to export",
                re.IGNORECASE,
            )
        ).first

        if await candidate.count() > 0 and await candidate.is_visible():
            dialog = candidate

    except Exception:
        pass

    # Fallback for Power BI overlay markup
    if dialog is None:
        try:
            candidate = page.locator(
                "[role='dialog'], .cdk-overlay-pane"
            ).filter(
                has_text=re.compile(
                    r"which data do you want to export",
                    re.IGNORECASE,
                )
            ).first

            if await candidate.count() > 0 and await candidate.is_visible():
                dialog = candidate

        except Exception:
            pass

    # ---------------------------------------------------------
    # Some Power BI versions download immediately.
    # ---------------------------------------------------------
    if dialog is None:
        logger.info(
            "Export dialog not displayed | "
            "Power BI appears to use direct export"
        )

        return {
            "data_type": "full",
            "option": "direct_export",
            "note": "Full data export successful.",
        }

    logger.info(
        "Power BI export-options dialog detected"
    )

    # ---------------------------------------------------------
    # OPTION 1: Data with current layout
    # ---------------------------------------------------------
    current_layout = dialog.get_by_text(
        re.compile(
            r"^\s*data with current layout\s*$",
            re.IGNORECASE,
        )
    ).first

    if await current_layout.count() > 0:
        logger.info(
            "Export option available: Data with current layout"
        )

        try:
            await current_layout.click(timeout=3000)
        except Exception:
            await current_layout.click(
                timeout=3000,
                force=True,
            )

        logger.info(
            "Selected export option: Data with current layout"
        )

        # Click final Export button
        export_button = dialog.get_by_role(
            "button",
            name=re.compile(
                r"^\s*export\s*$",
                re.IGNORECASE,
            ),
        ).first

        if await export_button.count() == 0:
            export_button = dialog.get_by_text(
                re.compile(
                    r"^\s*export\s*$",
                    re.IGNORECASE,
                )
            ).last

        if await export_button.count() > 0:
            try:
                await export_button.click(timeout=5000)
            except Exception:
                await export_button.click(
                    timeout=5000,
                    force=True,
                )

            logger.info(
                "Clicked final Export | option=Data with current layout"
            )

            return {
                "data_type": "full",
                "option": "Data with current layout",
                "note": "Full data export successful.",
            }

    logger.warning(
        "Data with current layout unavailable"
    )

    # ---------------------------------------------------------
    # OPTION 2: Summarized data
    # ---------------------------------------------------------
    summarized = dialog.get_by_text(
        re.compile(
            r"^\s*summarized data\s*$",
            re.IGNORECASE,
        )
    ).first

    if await summarized.count() > 0:
        logger.warning(
            "Falling back to Summarized data"
        )

        try:
            await summarized.click(timeout=3000)
        except Exception:
            await summarized.click(
                timeout=3000,
                force=True,
            )

        logger.warning(
            "Selected export option: Summarized data"
        )

        export_button = dialog.get_by_role(
            "button",
            name=re.compile(
                r"^\s*export\s*$",
                re.IGNORECASE,
            ),
        ).first

        if await export_button.count() == 0:
            export_button = dialog.get_by_text(
                re.compile(
                    r"^\s*export\s*$",
                    re.IGNORECASE,
                )
            ).last

        if await export_button.count() > 0:
            try:
                await export_button.click(timeout=5000)
            except Exception:
                await export_button.click(
                    timeout=5000,
                    force=True,
                )

            logger.warning(
                "Clicked final Export | "
                "validation will use SUMMARIZED DATA"
            )

            return {
                "data_type": "summarized",
                "option": "Summarized data",
                "note": (
                    "Full export unavailable; "
                    "using SUMMARIZED DATA."
                ),
            }

    raise RuntimeError(
        "Power BI export dialog was found, but neither "
        "'Data with current layout' nor 'Summarized data' "
        "could be selected."
    )

# ---------------------------------------------------------------------------
# Single visual export
# ---------------------------------------------------------------------------

async def _export_visual(
    page,
    visual: dict[str, Any],
    dashboard_name: str,
) -> dict[str, Any]:

    """
    Export ONE real Table/Matrix visual.

    Flow:

        find visual
            ↓
        close stale overlay
            ↓
        hover visual
            ↓
        More options
            ↓
        Export data
            ↓
        Export dialog
            ↓
        Data with current layout
            OR
        Summarized data
            ↓
        Final Export
            ↓
        download
            ↓
        parse
    """

    result = {
        "title": visual["title"],
        "visual_index": visual.get("index"),
        "status": "failed",
        "file_path": None,
        "data": None,
        "error": None,

        # How was this table validated?
        "validation_data_type": "unavailable",
        "validation_option": None,
        "validation_note": None,
    }

    last_error = None

    for attempt in range(
        1,
        MAX_EXPORT_RETRIES + 1,
    ):

        try:

            logger.info(
                "Export attempt %s/%s | "
                "dashboard=%s | visual=%s",
                attempt,
                MAX_EXPORT_RETRIES,
                dashboard_name,
                visual["title"],
            )

            # ---------------------------------------------------------
            # Open More options
            # ---------------------------------------------------------

            opened = await _open_more_options(
                page,
                visual,
            )

            if not opened:

                last_error = (
                    "Could not open More options."
                )

                continue

            # ---------------------------------------------------------
            # Find Export data
            # ---------------------------------------------------------

            item = await _find_export_data_item(
                page
            )

            if item is None:

                menu_text = ""

                try:

                    menu_text = await page.locator(
                        "[role='menu'], "
                        ".cdk-overlay-pane"
                    ).all_inner_texts()

                except Exception:
                    pass

                last_error = (
                    "Export data menu item was not found "
                    "after opening More options."
                )

                logger.warning(
                    "%s | menu_text=%s",
                    last_error,
                    menu_text,
                )

                await _close_open_overlays(
                    page
                )

                continue

            # ---------------------------------------------------------
            # Prepare download
            # ---------------------------------------------------------

            RAW_DIR.mkdir(
                parents=True,
                exist_ok=True,
            )

            async with page.expect_download(
                timeout=DOWNLOAD_TIMEOUT
            ) as download_info:

                logger.info(
                    "Clicking Export data | "
                    "dashboard=%s | visual=%s",
                    dashboard_name,
                    visual["title"],
                )

                try:

                    await item.click(
                        timeout=5000
                    )

                except Exception:

                    await item.click(
                        timeout=5000,
                        force=True,
                    )

                await page.wait_for_timeout(
                    500
                )

                # -----------------------------------------------------
                # Handle export options
                #
                # Priority:
                #   1. Data with current layout
                #   2. Summarized data
                # -----------------------------------------------------

                export_info = await _handle_export_dialog(
                    page
                )

            # ---------------------------------------------------------
            # Download completed
            # ---------------------------------------------------------

            download = await download_info.value

            suggested_name = (
                download.suggested_filename
            )

            suffix = (
                Path(
                    suggested_name
                ).suffix
                or ".csv"
            )

            filename = (
                f"{_safe_filename(dashboard_name, 'dashboard')}_"
                f"{_safe_filename(visual['title'], 'table')}_"
                f"{uuid.uuid4().hex[:8]}"
                f"{suffix}"
            )

            path = RAW_DIR / filename

            await download.save_as(
                str(path)
            )

            # ---------------------------------------------------------
            # Parse downloaded export
            # ---------------------------------------------------------

            data = _read_export(
                path
            )

            # ---------------------------------------------------------
            # Store result
            # ---------------------------------------------------------

            result.update(
                status="downloaded",
                file_path=str(path),
                data=data,
                validation_data_type=export_info.get(
                    "data_type",
                    "unknown",
                ),
                validation_option=export_info.get(
                    "option"
                ),
                validation_note=export_info.get(
                    "note"
                ),
            )

            # ---------------------------------------------------------
            # Explicit validation logging
            # ---------------------------------------------------------

            if (
                result["validation_data_type"]
                == "summarized"
            ):

                logger.warning(
                    "VALIDATION USING SUMMARIZED DATA | "
                    "dashboard=%s | visual=%s | "
                    "reason=%s",
                    dashboard_name,
                    visual["title"],
                    result["validation_note"],
                )

            elif (
                result["validation_data_type"]
                == "full"
            ):

                logger.info(
                    "VALIDATION USING FULL DATA | "
                    "dashboard=%s | visual=%s | "
                    "option=%s",
                    dashboard_name,
                    visual["title"],
                    result["validation_option"],
                )

            # ---------------------------------------------------------
            # Success log
            # ---------------------------------------------------------

            logger.info(
                "Export successful | "
                "dashboard=%s | visual=%s | "
                "rows=%s | columns=%s | "
                "validation_type=%s",
                dashboard_name,
                visual["title"],
                len(
                    data.get(
                        "rows",
                        [],
                    )
                ),
                len(
                    data.get(
                        "columns",
                        [],
                    )
                ),
                result["validation_data_type"],
            )

            # Close menu before next visual
            await _close_open_overlays(
                page
            )

            return result

        except Exception as exc:
            last_error = str(exc)

            logger.warning(
                "Export attempt %s failed | dashboard=%s | visual=%s | error=%s",
                attempt,
                dashboard_name,
                visual["title"],
                exc,
            )

            # If the browser is completely dead, stop trying to export.
            if "TargetClosedError" in str(exc) or "browser has been closed" in str(exc):
                logger.error("Browser closed unexpectedly. Aborting export for this visual.")
                break

            # Attempt graceful cleanup only if the browser is still alive
            try:
                await _close_open_overlays(page)
                await page.wait_for_timeout(750)
            except Exception as cleanup_exc:
                logger.debug("Cleanup failed: %s", cleanup_exc)
                if "TargetClosedError" in str(cleanup_exc) or "browser has been closed" in str(cleanup_exc):
                    break

    result["error"] = (
        last_error
        or "Export failed."
    )

    return result

async def export_table_visuals(
    page,
    table_visuals: list[dict[str, Any]],
    dashboard_name: str,
) -> list[dict[str, Any]]:
    """
    Export data for table/matrix visuals already identified by
    visual_data_exporter.py.

    This function does not:
    - launch a browser
    - navigate to a dashboard
    - discover all visuals
    - compare source and target tables

    It only exports and parses table/matrix data.
    """

    exported_tables: list[dict[str, Any]] = []

    if not table_visuals:
        logger.info(
            "No table/matrix visuals supplied for export | dashboard=%s",
            dashboard_name,
        )
        return exported_tables

    logger.info(
        "Starting table export | dashboard=%s | tables=%d",
        dashboard_name,
        len(table_visuals),
    )

    for table_number, table_visual in enumerate(
        table_visuals,
        start=1,
    ):
        try:
            logger.info(
                "Exporting table visual | dashboard=%s | number=%d/%d | "
                "title=%s | type=%s",
                dashboard_name,
                table_number,
                len(table_visuals),
                table_visual.get("title"),
                table_visual.get("visual_type"),
            )

            result = await _export_visual(
                page,
                table_visual,
                dashboard_name,
            )

            exported_tables.append(result)

            if result.get("status") == "downloaded":

                data = result.get("data") or {}

                logger.info(
                    "Table export successful | dashboard=%s | "
                    "title=%s | rows=%d | columns=%d",
                    dashboard_name,
                    result.get("title"),
                    len(data.get("rows", [])),
                    len(data.get("columns", [])),
                )

            else:

                logger.warning(
                    "Table export failed | dashboard=%s | "
                    "title=%s | error=%s",
                    dashboard_name,
                    table_visual.get("title"),
                    result.get("error"),
                )

        except Exception as exc:

            logger.exception(
                "Unexpected table export failure | dashboard=%s | "
                "title=%s",
                dashboard_name,
                table_visual.get("title"),
            )

            exported_tables.append(
                {
                    "title": table_visual.get("title"),
                    "visual_index": table_visual.get("index"),
                    "status": "failed",
                    "file_path": None,
                    "data": None,
                    "error": str(exc),
                    "validation_data_type": "unavailable",
                    "validation_option": None,
                    "validation_note": None,
                }
            )

    successful = sum(
        1
        for item in exported_tables
        if item.get("status") == "downloaded"
    )

    failed = len(exported_tables) - successful

    logger.info(
        "Table export completed | dashboard=%s | "
        "successful=%d | failed=%d",
        dashboard_name,
        successful,
        failed,
    )

    return exported_tables

# Add this at the bottom of table_exporter.py

from services.table_comparison import build_table_comparisons

def export_and_compare_tables(
    source_exports: list[dict[str, Any]], 
    target_exports: list[dict[str, Any]]
) -> dict[str, Any]:
    """
    Takes completed Source and Target export results, runs Pandas table comparison,
    and returns structured results including column mismatches for reporting services.
    """
    # 1. Format exports into the structure expected by build_table_comparisons()
    visual_data = {
        "Source": {
            "visuals": [
                {
                    "title": item.get("title"),
                    "data": item.get("data") or {"columns": [], "rows": []},
                    "scrollable": True
                }
                for item in source_exports if item.get("status") == "downloaded"
            ]
        },
        "Target": {
            "visuals": [
                {
                    "title": item.get("title"),
                    "data": item.get("data") or {"columns": [], "rows": []},
                    "scrollable": True
                }
                for item in target_exports if item.get("status") == "downloaded"
            ]
        }
    }

    # 2. Run the Pandas comparison engine
    comparison_results = build_table_comparisons(visual_data)

    # 3. Extract mismatched columns directly for Word document reporting
    column_mismatches = []
    for table in comparison_results.get("tables", []):
        source_only = table.get("source_only_columns", [])
        target_only = table.get("target_only_columns", [])
        col_diffs = table.get("column_differences", [])

        if source_only or target_only or col_diffs:
            column_mismatches.append({
                "visual": table.get("visual"),
                "status": table.get("status"),
                "source_only_columns": source_only,
                "target_only_columns": target_only,
                "column_differences": col_diffs,
            })

    # 4. Return combined comparison output
    return {
        **comparison_results,
        "column_mismatches": column_mismatches,
    }