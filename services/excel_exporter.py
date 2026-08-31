"""
Excel exporter for Browser Metrics Validator.

Creates one Excel workbook containing conditional sheets:
- Summary (Always created)
- Metadata
- Filters & Filter Comparison
- KPI Details & KPI Comparison
- Visual Details & Visual Comparison
- Browser Metrics & Network Details

Features:
- Dynamically creates sheets ONLY if data exists.
- Supports both AI extraction and DOM-only extraction payloads.
- Tabular-only visual data sheets (skips slicer/button visuals).
"""

import logging
import re
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)

# Ordered browser metrics sections aligned with automation/metrics.py
_BROWSER_METRIC_SECTIONS = [
    (
        "Run Information",
        [
            "validation_run_id",
            "run_id",
            "timestamp",
            "dashboard_name",
            "dashboard_url",
            "page_name",
            "page_title",
            "final_url",
            "http_status",
            "screenshot_path",
            "extraction_status",
            "extraction_error",
        ],
    ),
    (
        "Performance (seconds)",
        [
            "browser_launch_seconds",
            "page_load_seconds",
            "dashboard_render_seconds",
            "filter_dashboard_render_seconds",
            "screenshot_seconds",
            "gemini_extraction_seconds",
            "total_execution_seconds",
        ],
    ),
    (
        "Network Summary",
        [
            "total_requests",
            "total_responses",
            "failed_requests",
            "console_messages",
            "page_errors",
        ],
    ),
]

_SUMMARY_TIMING_METRICS = [
    ("Browser Launch (sec)", "browser_launch_seconds"),
    ("Page Load (sec)", "page_load_seconds"),
    ("Dashboard Render (sec)", "dashboard_render_seconds"),
    ("Filter Dashboard Render (sec)", "filter_dashboard_render_seconds"),
    ("Screenshot (sec)", "screenshot_seconds"),
    ("Gemini Extraction (sec)", "gemini_extraction_seconds"),
    ("Total Execution (sec)", "total_execution_seconds"),
    ("HTTP Requests", "total_requests"),
    ("Failed Requests", "failed_requests"),
    ("Console Messages", "console_messages"),
    ("Page Errors", "page_errors"),
]


def _kpi_key(name):
    """Canonicalise a KPI label without deleting business terms."""
    return " ".join(re.sub(r"[^a-z0-9%]+", " ", str(name or "").casefold()).split())


def _normalise_comparison_value(value):
    return re.sub(r"\s+", "", str(value or "")).replace(",", "").casefold()


def _match_abbreviated_kpis(source_map, target_map):
    matched = {}
    source_only = set(source_map) - set(target_map)
    target_only = set(target_map) - set(source_map)
    for target_key in target_only:
        target_tokens = set(target_key.split())
        target_value = _normalise_comparison_value(target_map[target_key].get("value"))
        candidates = [
            source_key for source_key in source_only
            if target_tokens and target_tokens < set(source_key.split())
            and target_value
            and target_value == _normalise_comparison_value(source_map[source_key].get("value"))
        ]
        if len(candidates) == 1:
            source_key = candidates[0]
            matched[source_key] = target_map[target_key]
            source_only.remove(source_key)
    return matched


# ---------------------------------------------------------
# Styles
# ---------------------------------------------------------

FONT_FAMILY = "Segoe UI"

HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SUBHEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True)
CELL_FONT = Font(name=FONT_FAMILY, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------
# Helper Functions
# ---------------------------------------------------------

def _format_header(ws, headers):
    ws.sheet_view.showGridLines = True
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def _style_cell(cell, alignment=ALIGN_LEFT):
    cell.font = CELL_FONT
    cell.border = THIN_BORDER
    cell.alignment = alignment


def _auto_fit(ws):
    for column_cells in ws.columns:
        if not column_cells:
            continue
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 4, 14), 60)


def _safe_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _values_to_string(value):
    if value is None:
        return "N/A"
    
    if isinstance(value, list):
        if not value:
            return "N/A"
            
        # Clean up DOM dictionaries into readable text
        cleaned_items = []
        seen_texts = set()
        
        for item in value:
            if isinstance(item, dict):
                # Pull the most readable text available from the DOM dict
                text = str(item.get('text') or item.get('title') or item.get('aria_label') or '').strip()
                
                # Skip pure noise and loading placeholders
                if not text or "loading" in text.lower() or "visual container" in text.lower() or "drill" in text.lower():
                    continue
                
                # Deduplicate so we don't repeat the same axis labels endlessly
                if text not in seen_texts:
                    seen_texts.add(text)
                    cleaned_items.append(text)
            else:
                cleaned_items.append(str(item))
                
        return ", ".join(cleaned_items) if cleaned_items else "N/A"
        
    return str(value)


# ---------------------------------------------------------
# Summary
# ---------------------------------------------------------

def _create_summary_sheet(wb, run_id, source_data, target_data, comparison_result, metrics):
    logger.info("Creating Summary worksheet")
    ws = wb.create_sheet("Summary")

    summary = comparison_result.get("summary", {})
    source_metadata = source_data.get("metadata", {})
    target_metadata = target_data.get("metadata", {})

    rows = [
        ("Run ID", run_id),
        ("Source Dashboard", source_metadata.get("dashboard_title") or "N/A"),
        ("Target Dashboard", target_metadata.get("dashboard_title") or "N/A"),
        ("Overall Match %", summary.get("overall_match_percentage", 0)),
        ("Filter Match %", summary.get("filter_match_percentage", 0)),
        ("KPI Match %", summary.get("kpi_match_percentage", 0)),
        ("Visual Match %", summary.get("visual_match_percentage", 0)),
        ("Source Filter Count", len(source_data.get("filters", []))),
        ("Target Filter Count", len(target_data.get("filters", []))),
        ("Source KPI Count", len(source_data.get("kpi_cards", []))),
        ("Target KPI Count", len(target_data.get("kpi_cards", []))),
        (
            "Source Visual Count",
            len(source_data.get("charts", []) or source_data.get("visuals", []))
            + len(source_data.get("tables", []) or source_data.get("table_exports", []))
            + len(source_data.get("kpi_cards", [])),
        ),
        (
            "Target Visual Count",
            len(target_data.get("charts", []) or target_data.get("visuals", []))
            + len(target_data.get("tables", []) or target_data.get("table_exports", []))
            + len(target_data.get("kpi_cards", [])),
        ),
    ]

    source_metrics = metrics[0] if metrics else None
    target_metrics = metrics[1] if metrics and len(metrics) > 1 else None
    source_label = source_metrics.get("dashboard_name", "Source") if source_metrics else "Source"
    target_label = target_metrics.get("dashboard_name", "Target") if target_metrics else "Target"

    for label, key in _SUMMARY_TIMING_METRICS:
        source_value = source_metrics.get(key, "N/A") if source_metrics else "N/A"
        target_value = target_metrics.get(key, "N/A") if target_metrics else "N/A"
        rows.append((label, f"{source_label}: {source_value} | {target_label}: {target_value}"))

    _format_header(ws, ["Property", "Value"])

    row_number = 2
    for property_name, value in rows:
        ws.cell(row=row_number, column=1, value=property_name)
        ws.cell(row=row_number, column=2, value=value)
        _style_cell(ws.cell(row=row_number, column=1))
        _style_cell(ws.cell(row=row_number, column=2), ALIGN_CENTER)
        row_number += 1

    _auto_fit(ws)


# ---------------------------------------------------------
# Metadata
# ---------------------------------------------------------

def _create_metadata_sheet(wb, source_data, target_data):
    logger.info("Creating Metadata worksheet")
    ws = wb.create_sheet("Metadata")

    _format_header(ws, ["Dashboard", "Property", "Value"])

    row = 2
    for dashboard_name, data in [("Source", source_data), ("Target", target_data)]:
        metadata = data.get("metadata", {})
        for key, value in metadata.items():
            ws.cell(row=row, column=1, value=dashboard_name)
            ws.cell(row=row, column=2, value=key.replace("_", " ").title())
            ws.cell(row=row, column=3, value=(str(value) if value is not None else "N/A"))
            for col in range(1, 4):
                _style_cell(ws.cell(row=row, column=col))
            row += 1

    _auto_fit(ws)


# ---------------------------------------------------------
# Filters
# ---------------------------------------------------------

def _create_filters_sheet(wb, source_data, target_data):
    logger.info("Creating Filters worksheet")
    ws = wb.create_sheet("Filters")

    headers = ["Dashboard", "Filter Name", "Filter Type", "Selected Values", "Available Values"]
    _format_header(ws, headers)

    row = 2
    for dashboard_name, data in [("Source", source_data), ("Target", target_data)]:
        filters = data.get("filters", [])
        for filter_data in filters:
            values = [
                dashboard_name,
                filter_data.get("filter_name", "N/A"),
                filter_data.get("filter_type", "N/A"),
                _values_to_string(filter_data.get("selected_values", [])),
                _values_to_string(filter_data.get("available_values", [])),
            ]
            for col_idx, value in enumerate(values, start=1):
                _style_cell(ws.cell(row=row, column=col_idx, value=value))
            row += 1

    _auto_fit(ws)


# ---------------------------------------------------------
# Filter Comparison
# ---------------------------------------------------------

def _create_filter_comparison_sheet(wb, comparison):
    logger.info("Creating Filter Comparison worksheet")
    ws = wb.create_sheet("Filter Comparison")

    headers = [
        "Filter Name", "Source Type", "Target Type",
        "Source Selected", "Target Selected",
        "Source Values", "Target Values", "Status"
    ]
    _format_header(ws, headers)

    row = 2
    for result in comparison:
        values = [
            result.get("filter_name", "N/A"),
            result.get("source_type", "N/A"),
            result.get("target_type", "N/A"),
            _values_to_string(result.get("source_selected", [])),
            _values_to_string(result.get("target_selected", [])),
            _values_to_string(result.get("source_values", [])),
            _values_to_string(result.get("target_values", [])),
            result.get("status", "Unknown"),
        ]
        for col_idx, value in enumerate(values, start=1):
            _style_cell(ws.cell(row=row, column=col_idx, value=value))
        row += 1

    _auto_fit(ws)


# ---------------------------------------------------------
# KPI Details
# ---------------------------------------------------------

def _create_kpi_details_sheet(wb, source_data, target_data):
    logger.info("Creating KPI Details worksheet")
    ws = wb.create_sheet("KPI Details")

    headers = [
        "Dashboard", "Visual ID", "Metric Name",
        "Current Value", "Prior Period Value",
        "Variance", "Confidence"
    ]
    _format_header(ws, headers)

    row = 2
    for dashboard_name, data in [("Source", source_data), ("Target", target_data)]:
        for kpi in data.get("kpi_cards", []):
            confidence = kpi.get("confidence")
            confidence_str = f"{confidence * 100:.1f}%" if confidence is not None else "N/A"

            values = [
                dashboard_name,
                kpi.get("visual_id", "N/A"),
                kpi.get("name", "N/A"),
                kpi.get("value", "N/A"),
                kpi.get("previous_value", "N/A"),
                kpi.get("variance", "N/A"),
                confidence_str,
            ]
            for col_idx, value in enumerate(values, start=1):
                alignment = ALIGN_RIGHT if col_idx in [4, 5, 6, 7] else ALIGN_LEFT
                _style_cell(ws.cell(row=row, column=col_idx, value=value), alignment)
            row += 1

    _auto_fit(ws)


# ---------------------------------------------------------
# KPI Comparison
# ---------------------------------------------------------
def compare_kpis(source_data: dict, target_data: dict) -> list:
    logger.info("Starting KPI comparison")
    try:
        source_kpis = source_data.get("kpi_cards", [])
        target_kpis = target_data.get("kpi_cards", [])

        source_map = {_kpi_key(kpi.get("name")): kpi for kpi in source_kpis if kpi.get("name")}
        target_map = {_kpi_key(kpi.get("name")): kpi for kpi in target_kpis if kpi.get("name")}

        for source_key, target_kpi in _match_abbreviated_kpis(source_map, target_map).items():
            for target_key, candidate in list(target_map.items()):
                if candidate is target_kpi:
                    del target_map[target_key]
                    break
            target_map[source_key] = target_kpi

        results = []
        all_names = sorted(set(source_map.keys()) | set(target_map.keys()))

        for name in all_names:
            source = source_map.get(name)
            target = target_map.get(name)

            if source is None:
                status = "Missing in Source"
            elif target is None:
                status = "Missing in Target"
            else:
                source_value = _normalise_comparison_value(source.get("value"))
                target_value = _normalise_comparison_value(target.get("value"))
                source_previous = _normalise_comparison_value(source.get("previous_value"))
                target_previous = _normalise_comparison_value(target.get("previous_value"))
                source_variance = _normalise_comparison_value(source.get("variance"))
                target_variance = _normalise_comparison_value(target.get("variance"))

                if (source_value == target_value and source_previous == target_previous and source_variance == target_variance):
                    status = "Match"
                elif source_value != target_value:
                    status = "Value Changed"
                elif source_previous != target_previous:
                    status = "Previous Value Changed"
                elif source_variance != target_variance:
                    status = "Variance Changed"
                else:
                    status = "Mismatch"

            results.append({
                "kpi": source.get("name") if source else target.get("name"),
                "source_kpi_name": source.get("name") if source else None,
                "target_kpi_name": target.get("name") if target else None,
                "source": source.get("value") if source else None,
                "target": target.get("value") if target else None,
                "source_prior": source.get("previous_value") if source else None,
                "target_prior": target.get("previous_value") if target else None,
                "source_variance": source.get("variance") if source else None,
                "target_variance": target.get("variance") if target else None,
                "status": status,
            })

        return results
    except Exception:
        logger.exception("Error during KPI comparison")
        raise


def _create_kpi_comparison_sheet(wb, comparison):
    logger.info("Creating KPI Comparison worksheet")
    ws = wb.create_sheet("KPI Comparison")

    headers = [
        "KPI", "Source Value", "Target Value",
        "Source Prior Value", "Target Prior Value",
        "Source Variance", "Target Variance", "Status"
    ]
    _format_header(ws, headers)

    row = 2
    for result in comparison:
        values = [
            result.get("kpi", "N/A"),
            result.get("source", "N/A"),
            result.get("target", "N/A"),
            result.get("source_prior", "N/A"),
            result.get("target_prior", "N/A"),
            result.get("source_variance", "N/A"),
            result.get("target_variance", "N/A"),
            result.get("status", "Unknown"),
        ]
        for col_idx, value in enumerate(values, start=1):
            _style_cell(ws.cell(row=row, column=col_idx, value=value))
        row += 1

    _auto_fit(ws)


# ---------------------------------------------------------
# Visual Details
# ---------------------------------------------------------

def _create_visual_details_sheet(wb, source_data, target_data, visual_comparison=None, table_comparison=None):
    logger.info("Creating Visual Details worksheet")
    ws = wb.create_sheet("Visual Details")

    headers = ["Dashboard", "Visual ID", "Visual Type", "Title", "Data", "Confidence", "Comparison Status"]
    _format_header(ws, headers)

    status_by_visual = {str(item.get("visual_id")): item.get("status", "Not Compared") for item in (visual_comparison or []) if item.get("visual_id")}
    table_status_by_title = {" ".join(str(item.get("visual", "")).casefold().split()): item.get("status", "Not Compared") for item in (table_comparison or {}).get("summary", [])}
    
    row = 2
    for dashboard_name, data in [("Source", source_data), ("Target", target_data)]:

        # Fallback to DOM visuals if AI charts are empty
        charts = data.get("charts", []) or data.get("visuals", [])
        for chart in charts:
            confidence = chart.get("confidence")
            confidence_str = f"{confidence * 100:.1f}%" if confidence is not None else "N/A"

            values = [
                dashboard_name,
                chart.get("visual_id", "N/A"),
                chart.get("chart_type") or chart.get("visual_type", "Chart"),
                chart.get("chart_title") or chart.get("title", "Untitled"),
                _values_to_string(chart.get("data") or chart.get("dom_content", [])),
                confidence_str,
                status_by_visual.get(str(chart.get("visual_id")), "Not Compared"),
            ]
            for col_idx, value in enumerate(values, start=1):
                _style_cell(ws.cell(row=row, column=col_idx, value=value))
            row += 1

        # Fallback to DOM table_exports if AI tables are empty
        tables = data.get("tables", []) or data.get("table_exports", [])
        for table in tables:
            title = table.get("table_title") or table.get("title", "Table")
            values = [
                dashboard_name,
                table.get("visual_id", "N/A"),
                "Table",
                title,
                f"{len(table.get('rows', []))} rows × {len(table.get('columns', []))} columns; see Table Data",
                "N/A",
                table_status_by_title.get(" ".join(str(title).casefold().split()), "Not Compared"),
            ]
            for col_idx, value in enumerate(values, start=1):
                _style_cell(ws.cell(row=row, column=col_idx, value=value))
            row += 1

    _auto_fit(ws)


# ---------------------------------------------------------
# Visual Comparison
# ---------------------------------------------------------

def _create_visual_comparison_sheet(wb, comparison):
    logger.info("Creating Visual Comparison worksheet")
    ws = wb.create_sheet("Visual Comparison")

    headers = ["Visual", "Source", "Target", "Status"]
    _format_header(ws, headers)

    row = 2
    for result in comparison:
        values = [
            result.get("visual", result.get("visual_id", "N/A")),
            _values_to_string(result.get("source", "N/A")),
            _values_to_string(result.get("target", "N/A")),
            result.get("status", "Unknown"),
        ]
        for col_idx, value in enumerate(values, start=1):
            _style_cell(ws.cell(row=row, column=col_idx, value=value))
        row += 1

    _auto_fit(ws)


# ---------------------------------------------------------
# Browser Metrics
# ---------------------------------------------------------

def _format_metric_value(value):
    if value is None:
        return "N/A"
    if isinstance(value, (list, dict)):
        return _values_to_string(value)
    return value


def _browser_metric_status(source_value, target_value):
    if source_value is None and target_value is None:
        return "N/A"
    if source_value is None:
        return "Missing in Source"
    if target_value is None:
        return "Missing in Target"

    source_text = str(_format_metric_value(source_value))
    target_text = str(_format_metric_value(target_value))
    if source_text == target_text:
        return "Match"

    try:
        if float(source_text) == float(target_text):
            return "Match"
    except (ValueError, TypeError):
        pass

    return "Different"


def _create_browser_metrics_sheet(wb, metrics):
    logger.info("Creating Browser Metrics worksheet")
    ws = wb.create_sheet("Browser Metrics")

    source_metrics = metrics[0] if metrics else None
    target_metrics = metrics[1] if metrics and len(metrics) > 1 else None
    source_label = source_metrics.get("dashboard_name", "Source") if source_metrics else "Source"
    target_label = target_metrics.get("dashboard_name", "Target") if target_metrics else "Target"

    headers = ["Category", "Metric", source_label, target_label]
    _format_header(ws, headers)

    row = 2
    section_keys = {key for _, keys in _BROWSER_METRIC_SECTIONS for key in keys}
    ordered_rows = []

    for category, keys in _BROWSER_METRIC_SECTIONS:
        for key in keys:
            if (source_metrics and key in source_metrics) or (target_metrics and key in target_metrics):
                ordered_rows.append((category, key))

    extra_keys = set()
    for dashboard_metrics in (source_metrics, target_metrics):
        if not dashboard_metrics:
            continue
        for key in dashboard_metrics:
            if key in section_keys or key in {"network_details", "dashboard_name"}:
                continue
            extra_keys.add(key)

    for key in sorted(extra_keys):
        ordered_rows.append(("Additional", key))

    for category, key in ordered_rows:
        source_value = source_metrics.get(key) if source_metrics else None
        target_value = target_metrics.get(key) if target_metrics else None
        values = [
            category, key,
            _format_metric_value(source_value),
            _format_metric_value(target_value)
        ]
        for col_idx, item in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=item)
            _style_cell(cell)
        row += 1

    if row == 2:
        for col_idx, value in enumerate(["N/A", "No browser metrics captured", "N/A", "N/A"], start=1):
            _style_cell(ws.cell(row=2, column=col_idx, value=value))

    _auto_fit(ws)


def _create_network_details_sheet(wb, metrics):
    logger.info("Creating Network Details worksheet")
    ws = wb.create_sheet("Network Details")

    headers = ["Dashboard", "Event Type", "Detail 1", "Detail 2", "Detail 3"]
    _format_header(ws, headers)

    row = 2
    for dashboard_metrics in metrics:
        if not dashboard_metrics:
            continue

        dashboard_name = dashboard_metrics.get("dashboard_name", "N/A")
        network_details = dashboard_metrics.get("network_details") or {}

        for item in network_details.get("failed_requests", []):
            values = [
                dashboard_name, "Failed Request",
                item.get("url", "N/A"), item.get("status", "N/A"), item.get("status_text", "N/A"),
            ]
            for col_idx, value in enumerate(values, start=1):
                _style_cell(ws.cell(row=row, column=col_idx, value=value))
            row += 1

        for item in network_details.get("console_logs", []):
            values = [
                dashboard_name, "Console",
                item.get("type", "N/A"), item.get("text", "N/A"), "",
            ]
            for col_idx, value in enumerate(values, start=1):
                _style_cell(ws.cell(row=row, column=col_idx, value=value))
            row += 1

        for item in network_details.get("page_errors", []):
            values = [
                dashboard_name, "Page Error", str(item), "", "",
            ]
            for col_idx, value in enumerate(values, start=1):
                _style_cell(ws.cell(row=row, column=col_idx, value=value))
            row += 1

    if row == 2:
        _style_cell(ws.cell(row=2, column=1, value="N/A"))
        _style_cell(ws.cell(row=2, column=2, value="No network events captured"))

    _auto_fit(ws)


def _visual_key(visual):
    return " ".join(str(visual.get("title") or visual.get("id") or "").lower().split())


def _column_key(value):
    return " ".join(str(value or "").casefold().split())


def _table_columns(visual):
    columns = visual.get("data", {}).get("columns", [])
    if columns:
        return [str(column) for column in columns]
    width = max((len(row) for row in visual.get("data", {}).get("rows", [])), default=0)
    return [f"Column {index}" for index in range(1, width + 1)]


def build_visual_data_comparison(visual_data):
    from services.table_comparison import is_tabular_visual

    source = [item for item in visual_data.get("Source", {}).get("visuals", []) if is_tabular_visual(item)]
    target = [item for item in visual_data.get("Target", {}).get("visuals", []) if is_tabular_visual(item)]
    
    source_map = {_visual_key(item): item for item in source if _visual_key(item)}
    target_map = {_visual_key(item): item for item in target if _visual_key(item)}
    
    summaries, cells = [], []
    for key in sorted(set(source_map) | set(target_map)):
        left, right = source_map.get(key), target_map.get(key)
        if not left or not right:
            summaries.append({"visual": (left or right).get("title"), "status": "Missing in Target" if left else "Missing in Source", "source_rows": len((left or {}).get("data", {}).get("rows", [])), "target_rows": len((right or {}).get("data", {}).get("rows", [])), "matched_cells": 0, "mismatched_cells": 0})
            continue
            
        left_rows, right_rows = left.get("data", {}).get("rows", []), right.get("data", {}).get("rows", [])
        left_columns, right_columns = _table_columns(left), _table_columns(right)
        
        if not left_rows and not right_rows and not left_columns and not right_columns:
            summaries.append({"visual": left.get("title"), "status": "No table data captured", "source_rows": 0, "target_rows": 0, "matched_cells": 0, "mismatched_cells": 0})
            continue
            
        left_column_map = {_column_key(name): index for index, name in enumerate(left_columns)}
        right_column_map = {_column_key(name): index for index, name in enumerate(right_columns)}
        column_keys = list(dict.fromkeys([_column_key(name) for name in left_columns + right_columns]))
        
        matched = mismatched = 0
        for row_number in range(max(len(left_rows), len(right_rows))):
            left_row = left_rows[row_number] if row_number < len(left_rows) else []
            right_row = right_rows[row_number] if row_number < len(right_rows) else []
            for column_key in column_keys:
                source_column = left_column_map.get(column_key)
                target_column = right_column_map.get(column_key)
                source_value = left_row[source_column] if source_column is not None and source_column < len(left_row) else None
                target_value = right_row[target_column] if target_column is not None and target_column < len(right_row) else None
                status = "Match" if source_value == target_value else ("Missing in Source" if source_value is None else "Missing in Target" if target_value is None else "Mismatch")
                matched += status == "Match"
                mismatched += status != "Match"
                display_column = left_columns[source_column] if source_column is not None else right_columns[target_column]
                cells.append({"visual": left.get("title"), "row_number": row_number + 1, "column": display_column, "source_value": source_value, "target_value": target_value, "status": status})
                
        summaries.append({"visual": left.get("title"), "status": "Match" if not mismatched else "Mismatch", "source_rows": len(left_rows), "target_rows": len(right_rows), "matched_cells": matched, "mismatched_cells": mismatched})
        
    return {"summary": summaries, "cells": cells}


def _create_visual_data_sheets(wb, visual_data):
    if not visual_data:
        return {"summary": [], "cells": []}
        
    comparison = build_visual_data_comparison(visual_data)
    
    # Conditional generation: only create sheets if there is actual table comparison data
    if not comparison.get("summary"):
        return comparison

    from services.table_comparison import is_tabular_visual

    summary = wb.create_sheet("Visual Data Summary")
    _format_header(summary, ["Visual", "Status", "Source Rows", "Target Rows", "Matched Cells", "Mismatched Cells"])
    for row_idx, item in enumerate(comparison["summary"], start=2):
        for col_idx, value in enumerate([item["visual"], item["status"], item["source_rows"], item["target_rows"], item["matched_cells"], item["mismatched_cells"]], start=1):
            _style_cell(summary.cell(row=row_idx, column=col_idx, value=value))
        if item["status"] != "Match":
            for col_idx in range(1, 7):
                summary.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    _auto_fit(summary)

    raw = wb.create_sheet("Table Data")
    source_visuals = {_visual_key(item): item for item in visual_data.get("Source", {}).get("visuals", []) if _visual_key(item) and is_tabular_visual(item)}
    target_visuals = {_visual_key(item): item for item in visual_data.get("Target", {}).get("visuals", []) if _visual_key(item) and is_tabular_visual(item)}
    
    row_idx = 1
    for key in sorted(set(source_visuals) | set(target_visuals)):
        source_visual, target_visual = source_visuals.get(key), target_visuals.get(key)
        source_columns = _table_columns(source_visual) if source_visual else []
        target_columns = _table_columns(target_visual) if target_visual else []
        source_width = max(1, len(source_columns))
        target_start = source_width + 3
        
        raw.cell(row=row_idx, column=1, value=f"Source: {(source_visual or target_visual).get('title')}")
        raw.cell(row=row_idx, column=target_start, value=f"Target: {(target_visual or source_visual).get('title')}")
        for cell in (raw.cell(row=row_idx, column=1), raw.cell(row=row_idx, column=target_start)):
            cell.font = SUBHEADER_FONT
        row_idx += 1
        
        for index, column in enumerate(source_columns, start=1):
            _style_cell(raw.cell(row=row_idx, column=index, value=column), ALIGN_CENTER)
            raw.cell(row=row_idx, column=index).font = HEADER_FONT
            raw.cell(row=row_idx, column=index).fill = HEADER_FILL
        for index, column in enumerate(target_columns, start=target_start):
            _style_cell(raw.cell(row=row_idx, column=index, value=column), ALIGN_CENTER)
            raw.cell(row=row_idx, column=index).font = HEADER_FONT
            raw.cell(row=row_idx, column=index).fill = HEADER_FILL
        row_idx += 1
        
        height = max(len((source_visual or {}).get("data", {}).get("rows", [])), len((target_visual or {}).get("data", {}).get("rows", [])))
        for data_row in range(height):
            for index, value in enumerate(((source_visual or {}).get("data", {}).get("rows", []) or [[]])[data_row] if data_row < len((source_visual or {}).get("data", {}).get("rows", [])) else [], start=1):
                _style_cell(raw.cell(row=row_idx, column=index, value=value))
            for index, value in enumerate(((target_visual or {}).get("data", {}).get("rows", []) or [[]])[data_row] if data_row < len((target_visual or {}).get("data", {}).get("rows", [])) else [], start=target_start):
                _style_cell(raw.cell(row=row_idx, column=index, value=value))
            row_idx += 1
        row_idx += 2
    _auto_fit(raw)

    cells = wb.create_sheet("Visual Data Comparison")
    _format_header(cells, ["Visual", "Row Number", "Column", "Source Value", "Target Value", "Status"])
    for row_idx, item in enumerate(comparison["cells"], start=2):
        for col_idx, value in enumerate([item["visual"], item["row_number"], item["column"], item["source_value"], item["target_value"], item["status"]], start=1):
            _style_cell(cells.cell(row=row_idx, column=col_idx, value=value))
        if item["status"] != "Match":
            for col_idx in range(1, 7):
                cells.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    _auto_fit(cells)
    
    return comparison


def _create_slicer_test_sheet(wb, scenarios):
    if not scenarios:
        return
    ws = wb.create_sheet("Slicer Test")
    _format_header(ws, ["Slicer", "Selected Value", "Source Applied", "Target Applied", "Visual", "Status", "Matching Cells", "Different Cells", "AI Analysis Error"])
    row = 2
    for scenario in scenarios:
        visual_results = scenario.get("visual_comparison") or [{}]
        for visual in visual_results:
            values = [scenario.get("slicer"), scenario.get("value"), scenario.get("source_applied"), scenario.get("target_applied"), visual.get("visual"), visual.get("status", scenario.get("status", "completed")), visual.get("matched_cells"), visual.get("mismatched_cells"), scenario.get("ai_analysis_error")]
            for col_idx, value in enumerate(values, start=1):
                _style_cell(ws.cell(row=row, column=col_idx, value=value))
            row += 1
    _auto_fit(ws)


# ---------------------------------------------------------
# Jagruthi — Arun multi-page dashboard sheets
# ---------------------------------------------------------

def _write_table_rows(ws, headers, rows, start_row=2):
    _format_header(ws, headers)
    row_number = start_row
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_number, column=col_idx, value=value)
            _style_cell(cell)
        row_number += 1
    _auto_fit(ws)
    return row_number


def _create_page_comparison_sheet(wb, page_comparisons):
    if not page_comparisons:
        return
    logger.info("Creating Page Comparison worksheet")
    ws = wb.create_sheet("Page Comparison")
    headers = ["Page Name", "Status", "Overall Match %", "Filter Match %", "KPI Match %", "Visual Match %"]
    rows = []
    for item in page_comparisons:
        summary = item.get("summary") or {}
        rows.append((item.get("page_name", "N/A"), item.get("status", "N/A"), summary.get("overall_match_percentage", 0), summary.get("filter_match_percentage", 0), summary.get("kpi_match_percentage", 0), summary.get("visual_match_percentage", 0)))
    _write_table_rows(ws, headers, rows)


def _create_page_inventory_sheet(wb, executions_by_dashboard):
    if not executions_by_dashboard:
        return
    from services.dashboard_inventory_service import _count_inventory_for_execution

    logger.info("Creating Page Inventory worksheet")
    ws = wb.create_sheet("Page Inventory")
    headers = ["Dashboard", "Page Name", "Filter Count", "KPI Count", "Table Count", "Matrix Count", "Chart Count", "Slicer Visuals", "Total Visuals", "Skipped Visuals"]
    rows = []
    for dashboard_executions in executions_by_dashboard:
        for execution in dashboard_executions:
            dashboard = execution.get("dashboard") or {}
            inventory = _count_inventory_for_execution(execution)
            rows.append((dashboard.get("name", "N/A"), dashboard.get("page_name") or inventory.get("page_name") or "Default", inventory.get("filter_count", 0), inventory.get("kpi_count", 0), inventory.get("table_count", 0), inventory.get("matrix_count", 0), inventory.get("chart_count", 0), inventory.get("slicer_visual_count", 0), inventory.get("total_visuals", 0), inventory.get("skipped_visual_count", 0)))
    _write_table_rows(ws, headers, rows)


def _create_page_kpis_sheet(wb, executions_by_dashboard):
    if not executions_by_dashboard:
        return
    from services.dashboard_inventory_service import _list_kpis_for_execution

    logger.info("Creating Page KPIs worksheet")
    ws = wb.create_sheet("Page KPIs")
    headers = ["Dashboard", "Page Name", "KPI", "Value", "Previous", "Variance", "Source"]
    rows = []
    for dashboard_executions in executions_by_dashboard:
        for execution in dashboard_executions:
            dashboard = execution.get("dashboard") or {}
            page_name = dashboard.get("page_name") or "Default"
            for kpi in _list_kpis_for_execution(execution):
                rows.append((dashboard.get("name", "N/A"), page_name, kpi.get("name"), kpi.get("value"), kpi.get("previous_value"), kpi.get("variance"), kpi.get("extraction_source")))
    _write_table_rows(ws, headers, rows)


def _create_page_visuals_sheet(wb, executions_by_dashboard):
    if not executions_by_dashboard:
        return
    from services.dashboard_inventory_service import _list_visuals_for_execution

    logger.info("Creating Page Visuals worksheet")
    ws = wb.create_sheet("Page Visuals")
    headers = ["Dashboard", "Page Name", "Visual Title", "Type", "Category", "Slicer", "Source"]
    rows = []
    for dashboard_executions in executions_by_dashboard:
        for execution in dashboard_executions:
            dashboard = execution.get("dashboard") or {}
            page_name = dashboard.get("page_name") or "Default"
            for visual in _list_visuals_for_execution(execution):
                rows.append((dashboard.get("name", "N/A"), page_name, visual.get("title"), visual.get("visual_type"), visual.get("category"), visual.get("is_slicer"), visual.get("extraction_source", "dom")))
    _write_table_rows(ws, headers, rows)


def _create_page_browser_metrics_sheet(wb, executions_by_dashboard):
    if not executions_by_dashboard:
        return

    logger.info("Creating Page Browser Metrics worksheet")
    ws = wb.create_sheet("Page Browser Metrics")
    headers = ["Dashboard", "Page Name", "Metric", "Value"]
    metric_keys = [key for _, keys in _BROWSER_METRIC_SECTIONS for key in keys]
    rows = []
    
    for dashboard_executions in executions_by_dashboard:
        for execution in dashboard_executions:
            dashboard = execution.get("dashboard") or {}
            metrics = execution.get("metrics") or {}
            page_name = dashboard.get("page_name") or metrics.get("page_name") or "Default"
            for key in metric_keys:
                if key not in metrics:
                    continue
                rows.append((dashboard.get("name", "N/A"), page_name, key, _format_metric_value(metrics.get(key))))
            for key in sorted(metrics):
                if key in metric_keys or key in {"network_details", "dashboard_name"}:
                    continue
                rows.append((dashboard.get("name", "N/A"), page_name, key, _format_metric_value(metrics.get(key))))
    _write_table_rows(ws, headers, rows)


# ---------------------------------------------------------
# Main Export Function
# ---------------------------------------------------------

def export_validation_workbook(
    run_id,
    source_data,
    target_data,
    filter_comparison,
    kpi_comparison,
    visual_comparison,
    comparison_summary,
    metrics,
    output_directory,
    visual_data=None,
    slicer_scenarios=None,
    page_comparisons=None,
    executions_by_dashboard=None,
):
    logger.info("Starting Excel workbook generation | run_id=%s", run_id)

    try:
        output_directory = Path(output_directory)
        output_directory.mkdir(parents=True, exist_ok=True)
        output_file = output_directory / f"{run_id}_dashboard_validation.xlsx"

        logger.info("Excel output path: %s", output_file)

        wb = Workbook()
        default_sheet = wb.active

        # -------------------------------------------------
        # Create all worksheets conditionally
        # -------------------------------------------------

        # 1. Summary & Metadata always exist
        _create_summary_sheet(wb, run_id, source_data, target_data, {"summary": comparison_summary}, metrics)
        _create_metadata_sheet(wb, source_data, target_data)

        # 2. Filters
        if source_data.get("filters") or target_data.get("filters"):
            _create_filters_sheet(wb, source_data, target_data)
        if filter_comparison:
            _create_filter_comparison_sheet(wb, filter_comparison)

        # 3. KPIs
        if source_data.get("kpi_cards") or target_data.get("kpi_cards"):
            _create_kpi_details_sheet(wb, source_data, target_data)
        if kpi_comparison:
            _create_kpi_comparison_sheet(wb, kpi_comparison)

        # 4. Visuals (Fallback logic applied here)
        charts_source = source_data.get("charts", []) or source_data.get("visuals", [])
        tables_source = source_data.get("tables", []) or source_data.get("table_exports", [])
        charts_target = target_data.get("charts", []) or target_data.get("visuals", [])
        tables_target = target_data.get("tables", []) or target_data.get("table_exports", [])

        if charts_source or tables_source or charts_target or tables_target:
            table_comparison = build_visual_data_comparison(visual_data or {})
            combined_visual_comparison = list(visual_comparison or []) + [
                {
                    "visual": item.get("visual"),
                    "source": f"{item.get('source_rows', 0)} rows / {item.get('matched_cells', 0)} matching cells",
                    "target": f"{item.get('target_rows', 0)} rows / {item.get('mismatched_cells', 0)} differing cells",
                    "status": item.get("status", "Not Compared"),
                }
                for item in table_comparison.get("summary", [])
            ]

            _create_visual_details_sheet(wb, source_data, target_data, visual_comparison, table_comparison)
            
            if combined_visual_comparison:
                _create_visual_comparison_sheet(wb, combined_visual_comparison)

        # 5. Metrics
        if metrics and any(metrics):
            _create_browser_metrics_sheet(wb, metrics)
            if any(m.get("network_details") for m in metrics if m):
                _create_network_details_sheet(wb, metrics)

        # 6. Arun multi-page dashboard data
        if page_comparisons or executions_by_dashboard:
            _create_page_comparison_sheet(wb, page_comparisons or [])
            _create_page_inventory_sheet(wb, executions_by_dashboard or [])
            _create_page_kpis_sheet(wb, executions_by_dashboard or [])
            _create_page_visuals_sheet(wb, executions_by_dashboard or [])
            _create_page_browser_metrics_sheet(wb, executions_by_dashboard or [])

        # 7. Deep Tabular Data (Self-checks if data exists before creating sheets)
        _create_visual_data_sheets(wb, visual_data)
        
        # 8. Slicer Tests
        if slicer_scenarios:
            _create_slicer_test_sheet(wb, slicer_scenarios)

        wb.remove(default_sheet)
        wb.save(output_file)

        logger.info("Excel workbook generated successfully | %s", output_file)
        return output_file

    except Exception:
        logger.exception("Failed to generate Excel workbook | run_id=%s", run_id)
        raise


def calculate_match_percentage(results: list) -> float | None:
    if not results:
        return None
    matches = sum(1 for result in results if result.get("status") == "Match")
    return round((matches / len(results)) * 100, 2)


def build_comparison_summary(filter_comparison, kpi_comparison, visual_comparison) -> dict:
    logger.info("Building dashboard comparison summary")
    
    filter_percentage = calculate_match_percentage(filter_comparison)
    kpi_percentage = calculate_match_percentage(kpi_comparison)
    visual_percentage = calculate_match_percentage(visual_comparison)

    scored = [value for value in (filter_percentage, kpi_percentage, visual_percentage) if value is not None]
    overall_percentage = round(sum(scored) / len(scored), 2) if scored else 0.0

    return {
        "filter_match_percentage": filter_percentage if filter_percentage is not None else 0.0,
        "kpi_match_percentage": kpi_percentage if kpi_percentage is not None else 0.0,
        "visual_match_percentage": visual_percentage if visual_percentage is not None else 0.0,
        "overall_match_percentage": overall_percentage,
        "kpi_compared": bool(kpi_comparison),
        "filter_compared": bool(filter_comparison),
        "visual_compared": bool(visual_comparison),
    }