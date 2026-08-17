"""Browser-side extraction of Power BI visual data.

Jagruthi — features:
- Null-safe slicer/filter DOM reads via page.evaluate
- Loading-placeholder detection and skip during visual scan
- Slicer vs table separation (button filters are not scraped as tables)
- DOM KPI card/callout extraction to supplement Gemini
- Vertical and horizontal 2D matrix scroll with merged column headers (Jagruthi)
"""

from __future__ import annotations

import asyncio
import csv
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


VISUAL_SELECTOR = ".visualContainer, [data-visual-container]"
_LOADING_TITLE_RE = re.compile(r"\bvisuals?\s+are\s+loading\b", re.IGNORECASE)
logger = logging.getLogger(__name__)


def _normalise_lines(value: str) -> list[str]:
    return [line.strip() for line in value.splitlines() if line.strip()]


def _safe_filename(value: str, fallback: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_.")
    return value or fallback


class _MatrixCellMerger:
    """Merge matrix cells across vertical and horizontal scroll positions."""

    def __init__(self) -> None:
        self.header_map: dict[str, str] = {}
        self.col_order: list[str] = []
        self.row_cells: dict[str, dict[str, str]] = {}

    @staticmethod
    def _col_key(item: dict[str, Any]) -> str:
        colindex = item.get("colindex")
        if colindex is not None and str(colindex).strip():
            return f"col:{colindex}"
        return f"left:{int(round(float(item.get("left", 0)) / 4)) * 4}"

    @staticmethod
    def _row_key(item: dict[str, Any]) -> str:
        rowindex = item.get("rowindex")
        if rowindex is not None and str(rowindex).strip():
            return f"row:{rowindex}"
        return f"top:{int(round(float(item.get("top", 0)) / 4)) * 4}"

    def ingest(self, payload: dict[str, Any] | None) -> None:
        if not payload:
            return
        for header in payload.get("headers", []):
            key = self._col_key(header)
            if header.get("value"):
                self.header_map.setdefault(key, header["value"])
            if key not in self.col_order:
                self.col_order.append(key)
        for cell in payload.get("cells", []):
            if cell.get("role") == "columnheader":
                key = self._col_key(cell)
                if cell.get("value"):
                    self.header_map.setdefault(key, cell["value"])
                if key not in self.col_order:
                    self.col_order.append(key)
                continue
            row_key = self._row_key(cell)
            col_key = self._col_key(cell)
            if col_key not in self.col_order:
                self.col_order.append(col_key)
            self.row_cells.setdefault(row_key, {})[col_key] = cell["value"]

    def row_count(self) -> int:
        return len(self.row_cells)

    def to_table(self) -> tuple[list[str], list[list[str]]]:
        if not self.col_order and not self.row_cells:
            return [], []
        columns = [
            self.header_map.get(key) or f"Column {index + 1}"
            for index, key in enumerate(self.col_order)
        ]

        def sort_key(key: str) -> tuple[int | str, ...]:
            if key.startswith("row:"):
                try:
                    return (0, int(key.split(":", 1)[1]))
                except ValueError:
                    return (1, key)
            if key.startswith("top:"):
                try:
                    return (0, int(key.split(":", 1)[1]))
                except ValueError:
                    return (1, key)
            return (1, key)

        rows = []
        for row_key in sorted(self.row_cells.keys(), key=sort_key):
            row_data = self.row_cells[row_key]
            rows.append([row_data.get(col_key, "") for col_key in self.col_order])
        return columns, rows


class VisualDataExporter:
    """Extract visible visual data without modifying the dashboard's filters."""

    def __init__(
        self,
        page,
        *,
        download_directory: str | Path | None = None,
        max_scroll_steps: int = 30,
    ):
        self.page = page
        self.download_directory = Path(download_directory) if download_directory else None
        self.max_scroll_steps = max_scroll_steps

    async def extract_dashboard_data(self, *, attempt_export: bool = False) -> dict[str, Any]:
        """Return slicer state, visual metadata, visible/scrollable rows and exports.

        ``attempt_export`` is opt-in because Power BI tenants can disable Export
        data and because clicking the visual menu changes the live UI.  DOM data
        collection remains available regardless of that setting.
        """
        result: dict[str, Any] = {
            "status": "success",
            "extracted_at": datetime.now(timezone.utc).isoformat(),
            "filters": await self._extract_filter_state(),
            "kpi_cards": await self._extract_kpi_cards(),
            "visuals": [],
            "skipped_visuals": [],
            "errors": [],
        }

        try:
            visual_count = await self.page.locator(VISUAL_SELECTOR).count()
        except Exception as exc:
            result.update(status="failed")
            result["errors"].append(f"Unable to locate Power BI visuals: {exc}")
            return result

        if not visual_count:
            result.update(status="partial")
            result["errors"].append("No Power BI visual containers were found.")
            return result

        for index in range(visual_count):
            locator = self.page.locator(VISUAL_SELECTOR).nth(index)
            try:
                if not await locator.is_visible():
                    result["skipped_visuals"].append({"index": index + 1, "reason": "hidden"})
                    continue
                visual = await self._inspect_visual(locator, index)
                # Jagruthi: skip Power BI loading shells before table/slicer handling.
                if visual.pop("is_loading_placeholder", False) or _LOADING_TITLE_RE.search(
                    visual.get("title", "")
                ):
                    result["skipped_visuals"].append({
                        "index": index + 1,
                        "reason": "Power BI loading placeholder",
                        "title": visual.get("title"),
                    })
                    continue
                if visual.get("is_slicer"):
                    visual["data"] = {
                        "columns": [],
                        "rows": [],
                        "row_count": 0,
                        "collection_method": "slicer_skipped",
                    }
                    result["skipped_visuals"].append({
                        "index": index + 1,
                        "reason": "slicer_or_button_filter",
                        "title": visual.get("title"),
                    })
                    continue
                # Native export is the accuracy fallback for virtualised Power
                # BI tables/matrices: request it only for visuals that expose
                # a grid/scrollbar, never for every chart on the page.
                is_tabular = bool(
                    visual["data"]["columns"]
                    or visual["scrollable"]
                    or visual["horizontally_scrollable"]
                    or (
                        visual["data"]["rows"]
                        and max(len(row) for row in visual["data"]["rows"]) > 1
                    )
                )
                if attempt_export and is_tabular and visual["export"]["supported"]:
                    visual["export"] = await self._try_export(locator, visual, index)
                    if visual["export"].get("status") == "downloaded":
                        exported = await asyncio.to_thread(self._read_exported_table, visual["export"]["file_path"])
                        if exported and exported["rows"]:
                            visual["data"] = exported
                            visual["data"]["collection_method"] = "power_bi_export"
                result["visuals"].append(visual)
            except Exception as exc:
                result["status"] = "partial"
                result["errors"].append(f"Visual {index + 1}: {exc}")

        logger.info("Visual extraction completed | visuals=%d | skipped=%d | errors=%d", len(result["visuals"]), len(result["skipped_visuals"]), len(result["errors"]))
        return result

    async def _extract_filter_state(self) -> list[dict[str, Any]]:
        """Read filters/buttons from the live Power BI accessibility DOM.

        Button slicers often do not expose the ``slicerContainer`` class, so
        examine visual containers as well and use their pressed/checked state
        instead of asking Gemini to infer selection from colour.
        """
        try:
            # Jagruthi: use page.evaluate with null-safe reads; avoid evaluate_all on broad selectors.
            filters = await self.page.evaluate(
                """() => {
                    const safeText = item => {
                        if (!item) return '';
                        return (item.innerText || item.getAttribute('aria-label') || item.value || '')
                            .replace(/\\s+/g, ' ').trim();
                    };
                    const nodeSelector = '.slicerContainer, [aria-label*="Slicer" i], [data-visual-type*="slicer" i]';
                    const nodes = [...document.querySelectorAll(nodeSelector)];
                    if (!nodes.length) {
                        return [...document.querySelectorAll('.visualContainer, [data-visual-container]')]
                            .map((visual, index) => {
                                const controls = [...visual.querySelectorAll(
                                    '[role="option"], [role="radio"], [role="checkbox"], label, button, input'
                                )]
                                    .filter(item => item && !item.disabled)
                                    .map(item => ({
                                        value: safeText(item),
                                        selected: item.matches(
                                            'input:checked, [aria-selected="true"], [aria-checked="true"], [aria-pressed="true"]'
                                        ) || /(^|\\s)(selected|active)(\\s|$)/i.test(item.className || ''),
                                        control_type: item.tagName === 'BUTTON' || item.getAttribute('role') === 'button'
                                            ? 'button'
                                            : (item.matches('input[type="radio"], input[type="checkbox"]') ? 'input' : 'option'),
                                    }))
                                    .filter(item => item.value);
                                const title = safeText(
                                    visual.querySelector('.visualTitle, [class*="visualTitle"], [data-visual-title]')
                                ) || safeText(visual);
                                const hasChoiceControl = controls.some(
                                    item => item.value && !/^(more options|focus mode|drill down|expand)$/i.test(item.value)
                                );
                                if (!title || !hasChoiceControl || controls.length < 2) return null;
                                const selected = controls.filter(item => item.selected).map(item => item.value);
                                const values = controls.map(item => item.value);
                                const hasButtons = controls.some(item => item.control_type === 'button');
                                return {
                                    id: visual.id || `filter-visual-${index + 1}`,
                                    name: title,
                                    filter_type: hasButtons ? 'Buttons' : 'Dropdown',
                                    selected_values: [...new Set(selected)],
                                    visible_values: [...new Set(values)].slice(0, 500),
                                    options: controls,
                                    looks_like_filter: true,
                                    extraction_source: 'dom',
                                };
                            })
                            .filter(Boolean);
                    }
                    return nodes.map((node, index) => {
                        const visual = node.closest('.visualContainer, [data-visual-container]') || node;
                        const title = safeText(
                            visual.querySelector('.visualTitle, [class*="visualTitle"], [data-visual-title]')
                        )
                            || (visual.getAttribute('aria-label') || '').trim()
                            || safeText(visual.querySelector('[title]'));
                        const controls = [...node.querySelectorAll(
                            '[role="option"], [role="radio"], [role="checkbox"], label, button, input'
                        )]
                            .filter(item => item && !item.disabled)
                            .map(item => ({
                                value: safeText(item),
                                selected: item.matches(
                                    'input:checked, [aria-selected="true"], [aria-checked="true"], [aria-pressed="true"]'
                                ) || /(^|\\s)(selected|active)(\\s|$)/i.test(item.className || ''),
                                control_type: item.tagName === 'BUTTON' || item.getAttribute('role') === 'button'
                                    ? 'button'
                                    : (item.matches('input[type="radio"], input[type="checkbox"]') ? 'input' : 'option'),
                            }))
                            .filter(item => item.value);
                        const selected = controls.filter(item => item.selected).map(item => item.value);
                        const values = controls.map(item => item.value);
                        const hasButtons = controls.some(item => item.control_type === 'button');
                        const hasSlicerMarkup = node.matches(
                            '.slicerContainer, [class*="slicer" i], [aria-label*="Slicer" i], [data-visual-type*="slicer" i]'
                        ) || visual.matches('[data-visual-type*="slicer" i]');
                        const hasChoiceControl = controls.some(
                            item => item.value && !/^(more options|focus mode|drill down|expand)$/i.test(item.value)
                        );
                        const looksLikeFilter = hasSlicerMarkup || (title && hasChoiceControl && controls.length >= 2);
                        return {
                            id: node.id || `slicer-${index + 1}`,
                            name: title,
                            filter_type: hasButtons ? 'Buttons' : (controls.some(item => item.selected) ? 'Buttons' : 'Dropdown'),
                            selected_values: [...new Set(selected)],
                            visible_values: [...new Set(values)].slice(0, 500),
                            options: controls,
                            looks_like_filter: looksLikeFilter,
                            extraction_source: 'dom',
                        };
                    }).filter(item => item.looks_like_filter && item.name);
                }"""
            )
            # The selector can overlap (a slicer is also a visual container).
            # Keep the record with the most actual option/selection evidence,
            # not simply the last wrapper returned by the DOM query.
            unique = {}
            for item in filters:
                key = " ".join(item["name"].casefold().split())
                score = len(item.get("visible_values", [])) + (1000 if item.get("selected_values") else 0)
                previous = unique.get(key)
                previous_score = len(previous.get("visible_values", [])) + (1000 if previous.get("selected_values") else 0) if previous else -1
                if score > previous_score:
                    unique[key] = item
            return list(unique.values())
        except Exception:
            logger.exception("Unable to read slicer state")
            return []

    async def _extract_kpi_cards(self) -> list[dict[str, Any]]:
        """Read compact KPI/card visuals from the DOM when Gemini misses them."""
        try:
            return await self.page.evaluate(
                """() => {
                    const safeText = item => {
                        if (!item) return '';
                        return (item.innerText || item.getAttribute('aria-label') || item.value || '')
                            .replace(/\\s+/g, ' ').trim();
                    };
                    const cards = [];
                    const visuals = [...document.querySelectorAll('.visualContainer, [data-visual-container]')];
                    for (const visual of visuals) {
                        const title = safeText(
                            visual.querySelector('.visualTitle, [class*="visualTitle"], [data-visual-title]')
                        ) || safeText(visual.querySelector('[title]'));
                        const valueNode = visual.querySelector(
                            '[class*="callout" i] [class*="value" i], [class*="calloutValue" i], [class*="dataLabel" i], [class*="metric" i], [class*="value" i]'
                        );
                        const value = safeText(valueNode);
                        if (!title || !value) continue;
                        if (/^(more options|focus mode|drill down|expand)$/i.test(title)) continue;
                        if (!/[\\d%,.$]/i.test(value)) continue;
                        const typeSource = [
                            visual.className,
                            visual.getAttribute('data-visual-type'),
                            visual.getAttribute('aria-roledescription'),
                        ].filter(Boolean).join(' ');
                        if (/slicer|dropdown|button/i.test(typeSource)) continue;
                        cards.push({
                            name: title,
                            value,
                            previous_value: null,
                            variance: null,
                            extraction_source: 'dom',
                        });
                    }
                    return cards;
                }"""
            )
        except Exception:
            logger.exception("Unable to read DOM KPI cards")
            return []

    async def _inspect_visual(self, locator, index: int) -> dict[str, Any]:
        metadata = await locator.evaluate(
            r"""(node, index) => {
                const text = (node.innerText || '').trim();
                const titleNode = node.querySelector('[title], [aria-label], .title, .visualTitle');
                const typeSource = [node.className, node.getAttribute('data-visual-type'), node.getAttribute('aria-roledescription')]
                    .filter(Boolean).join(' ');
                const canScrollY = (el) => el.scrollHeight > el.clientHeight + 2;
                const canScrollX = (el) => el.scrollWidth > el.clientWidth + 2;
                const scrollable = [...node.querySelectorAll('*')].some(el => {
                    const style = getComputedStyle(el);
                    return canScrollY(el) && /(auto|scroll|hidden)/i.test(style.overflowY);
                });
                const horizontallyScrollable = [...node.querySelectorAll('*')].some(el => {
                    const style = getComputedStyle(el);
                    return canScrollX(el) && /(auto|scroll|hidden)/i.test(style.overflowX);
                });
                const exportLabel = [...node.querySelectorAll('button, [role="button"]')]
                    .map(el => `${el.getAttribute('aria-label') || ''} ${el.title || ''} ${el.innerText || ''}`)
                    .some(text => /export data|more options|more/i.test(text));
                const isSlicer = /slicer/i.test(typeSource)
                    || node.matches('.slicerContainer, [class*="slicer" i], [aria-label*="Slicer" i], [data-visual-type*="slicer" i]')
                    || node.querySelector('.slicerContainer, [class*="slicer" i], [aria-label*="Slicer" i]');
                return {
                    id: node.getAttribute('data-visual-id') || node.id || `visual-${index + 1}`,
                    title: (titleNode?.getAttribute('title') || titleNode?.getAttribute('aria-label') || '').trim(),
                    visual_type: typeSource || 'unknown',
                    accessible_text: text,
                    is_loading_placeholder: /\bvisuals?\s+are\s+loading\b/i.test(text) || /^loading\.\.\.?$/i.test(text),
                    is_slicer: Boolean(isSlicer),
                    scrollable,
                    horizontally_scrollable: horizontallyScrollable,
                    export_menu_present: exportLabel,
                };
            }""",
            index,
        )
        rows: list[list[str]] = []
        columns: list[str] = []
        scanned_columns = 0
        if not metadata.get("is_slicer"):
            columns, rows, scanned_columns = await self._collect_rows(
                locator,
                metadata.get("scrollable", False),
                metadata.get("horizontally_scrollable", False),
            )
        if not columns:
            columns = await locator.evaluate(
                """node => [...node.querySelectorAll('[role="columnheader"], th, .columnHeader, [class*="columnHeader" i]')]
                    .map(cell => (cell.innerText || cell.getAttribute('aria-label') || '').trim())
                    .filter(Boolean)"""
            )
        if columns and rows and rows[0] == columns:
            rows = rows[1:]
        lines = _normalise_lines(metadata.pop("accessible_text", ""))
        title = metadata["title"] or (lines[0] if lines else metadata["id"])
        metadata["title"] = title
        if _LOADING_TITLE_RE.search(title) or _LOADING_TITLE_RE.search(metadata.get("title", "")):
            metadata["is_loading_placeholder"] = True
        metadata["data"] = {
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "scanned_columns": scanned_columns,
            "collection_method": "rendered_dom",
        }
        if not metadata.get("is_loading_placeholder"):
            logger.info(
                "Visual collected | title=%s | rows=%d | cols=%d | vertical_scroll=%s | horizontal_scroll=%s",
                title,
                len(rows),
                len(columns),
                metadata["scrollable"],
                metadata["horizontally_scrollable"],
            )
        metadata["export"] = {
            "supported": metadata.pop("export_menu_present"),
            "attempted": False,
            "status": "not_attempted",
            "file_path": None,
        }
        return metadata

    async def _collect_rows(
        self,
        locator,
        scrollable: bool,
        horizontally_scrollable: bool,
    ) -> tuple[list[str], list[list[str]], int]:
        """Collect matrix/table cells with combined vertical and horizontal scrolling.

        Jagruthi: Power BI matrices virtualise both axes.  Cells are merged by
        aria-rowindex/colindex (or position) so new columns from horizontal
        scroll extend existing rows instead of creating duplicate partial rows.
        """
        merger = _MatrixCellMerger()

        async def snapshot() -> None:
            payload = await locator.evaluate(
                """node => {
                    const text = el => {
                        if (!el) return '';
                        return (el.innerText || el.getAttribute('aria-label') || '').trim();
                    };
                    const headers = [];
                    const cells = [];
                    node.querySelectorAll(
                        '[role="columnheader"], th, .columnHeader, [class*="columnHeader" i]'
                    ).forEach(cell => {
                        const value = text(cell);
                        const box = cell.getBoundingClientRect();
                        if (!value || box.width < 1) return;
                        headers.push({
                            value,
                            colindex: cell.getAttribute('aria-colindex'),
                            left: box.left,
                        });
                    });
                    const selectors = '[role="gridcell"], [role="rowheader"], [role="columnheader"], td, th';
                    node.querySelectorAll(selectors).forEach(cell => {
                        const value = text(cell);
                        if (!value) return;
                        const box = cell.getBoundingClientRect();
                        if (box.width < 1 || box.height < 1) return;
                        cells.push({
                            value,
                            rowindex: cell.getAttribute('aria-rowindex'),
                            colindex: cell.getAttribute('aria-colindex'),
                            left: box.left,
                            top: box.top,
                            role: cell.getAttribute('role') || '',
                        });
                    });
                    if (!cells.length) {
                        [...node.querySelectorAll('[role="row"], tr')].forEach(row => {
                            const rowCells = [...row.querySelectorAll(
                                '[role="gridcell"], [role="columnheader"], td, th, .cell, [class*="cell" i]'
                            )].map(cell => text(cell)).filter(Boolean);
                            if (!rowCells.length) return;
                            rowCells.forEach((value, index) => {
                                cells.push({
                                    value,
                                    rowindex: null,
                                    colindex: String(index + 1),
                                    left: index * 100,
                                    top: row.offsetTop || 0,
                                    role: 'gridcell',
                                });
                            });
                        });
                    }
                    return { headers, cells };
                }"""
            )
            merger.ingest(payload)

        async def reset_axis(axis: str) -> None:
            await locator.evaluate(
                """(node, axis) => {
                    const isY = axis === 'y';
                    [...node.querySelectorAll('*')].forEach(el => {
                        const size = isY ? el.scrollHeight - el.clientHeight : el.scrollWidth - el.clientWidth;
                        if (size > 2) {
                            if (isY) el.scrollTop = 0;
                            else el.scrollLeft = 0;
                        }
                    });
                }""",
                axis,
            )

        async def scroll_axis(axis: str) -> bool:
            return await locator.evaluate(
                """(node, axis) => {
                    const isY = axis === 'y';
                    const candidates = [...node.querySelectorAll('*')].filter(el => {
                        const delta = isY
                            ? el.scrollHeight - el.clientHeight
                            : el.scrollWidth - el.clientWidth;
                        return delta > 2;
                    }).sort((a, b) => {
                        const aDelta = isY ? a.scrollHeight - a.clientHeight : a.scrollWidth - a.clientWidth;
                        const bDelta = isY ? b.scrollHeight - b.clientHeight : b.scrollWidth - b.clientWidth;
                        return bDelta - aDelta;
                    });
                    const element = candidates[0];
                    if (!element) return false;
                    const before = isY ? element.scrollTop : element.scrollLeft;
                    const max = isY
                        ? element.scrollHeight - element.clientHeight
                        : element.scrollWidth - element.clientWidth;
                    const step = Math.max((isY ? element.clientHeight : element.clientWidth) * 0.75, 40);
                    if (isY) {
                        element.scrollTop = Math.min(element.scrollTop + step, max);
                    } else {
                        element.scrollLeft = Math.min(element.scrollLeft + step, max);
                    }
                    element.dispatchEvent(new Event('scroll', { bubbles: true }));
                    const after = isY ? element.scrollTop : element.scrollLeft;
                    return after > before + 0.5;
                }""",
                axis,
            )

        await snapshot()
        horizontal_moves = 0
        vertical_moves = 0

        for _ in range(self.max_scroll_steps + 1):
            await reset_axis("y")
            await snapshot()

            if scrollable or merger.row_count() > 0:
                for _ in range(self.max_scroll_steps):
                    moved = await scroll_axis("y")
                    if not moved:
                        break
                    vertical_moves += 1
                    await self.page.wait_for_timeout(250)
                    await snapshot()

            if not horizontally_scrollable and horizontal_moves == 0:
                probe = await scroll_axis("x")
                if probe:
                    horizontal_moves += 1
                    horizontally_scrollable = True
                    await self.page.wait_for_timeout(250)
                    await reset_axis("y")
                    await snapshot()
                    continue
                break

            moved_right = await scroll_axis("x")
            if not moved_right:
                break
            horizontal_moves += 1
            await self.page.wait_for_timeout(250)

        columns, rows = merger.to_table()
        logger.info(
            "Matrix scan complete | vertical_steps=%d | horizontal_steps=%d | rows=%d | cols=%d",
            vertical_moves,
            horizontal_moves,
            len(rows),
            len(columns),
        )
        return columns, rows, len(columns)

    async def _try_export(self, locator, visual: dict[str, Any], index: int) -> dict[str, Any]:
        export = dict(visual["export"], attempted=True, status="unavailable")
        try:
            menu = locator.locator("button, [role='button']").filter(has_text=re.compile("more options|more", re.I))
            if await menu.count() == 0:
                return export
            await menu.first.click()
            item = self.page.get_by_text(re.compile(r"export data", re.I)).first
            if await item.count() == 0:
                return export
            async with self.page.expect_download(timeout=10_000) as download_info:
                await item.click()
            download = await download_info.value
            if self.download_directory:
                self.download_directory.mkdir(parents=True, exist_ok=True)
                path = self.download_directory / _safe_filename(visual["title"], f"visual-{index + 1}")
                path = path.with_suffix(Path(download.suggested_filename).suffix or ".csv")
                await download.save_as(str(path))
                export["file_path"] = str(path)
            export["status"] = "downloaded"
        except Exception as exc:
            export["error"] = str(exc)
        return export

    @staticmethod
    def _read_exported_table(file_path: str | None) -> dict[str, Any] | None:
        """Read a Power BI CSV/XLSX export into the common table schema."""
        if not file_path:
            return None
        path = Path(file_path)
        try:
            if path.suffix.lower() == ".csv":
                with path.open("r", encoding="utf-8-sig", newline="") as handle:
                    values = list(csv.reader(handle))
            elif path.suffix.lower() in {".xlsx", ".xlsm"}:
                workbook = load_workbook(path, read_only=True, data_only=True)
                sheet = workbook.active
                values = [list(row) for row in sheet.iter_rows(values_only=True)]
                workbook.close()
            else:
                logger.warning("Unsupported Power BI export format | %s", path.suffix)
                return None
            values = [["" if value is None else str(value) for value in row] for row in values if any(value is not None and str(value).strip() for value in row)]
            if not values:
                return None
            return {"columns": values[0], "rows": values[1:], "row_count": max(0, len(values) - 1)}
        except Exception:
            logger.exception("Unable to parse Power BI export | %s", path)
            return None


async def extract_visual_data(page, **kwargs) -> dict[str, Any]:
    """Convenience API used by the validator and future API/frontend callers."""
    attempt_export = kwargs.pop("attempt_export", False)
    return await VisualDataExporter(page, **kwargs).extract_dashboard_data(
        attempt_export=attempt_export
    )


async def extract_filter_data(page) -> dict[str, Any]:
    """Collect slicer/filter state only — skips table/matrix scrolling and exports."""
    exporter = VisualDataExporter(page)
    try:
        filters = await exporter._extract_filter_state()
        return {
            "status": "success",
            "extracted_at": datetime.now(timezone.utc).isoformat(),
            "filters": filters,
            "errors": [],
        }
    except Exception as exc:
        logger.exception("Filter-only extraction failed")
        return {
            "status": "failed",
            "extracted_at": datetime.now(timezone.utc).isoformat(),
            "filters": [],
            "errors": [str(exc)],
        }


async def apply_slicer_value(page, slicer_name: str, value: str) -> bool:
    """Select an exact slicer option in the currently loaded Power BI page.

    Returns ``False`` when the accessible DOM does not expose a safe matching
    control; callers can report this rather than claiming a filter was applied.
    """
    try:
        # Jagruthi: null-safe slicer click via page.evaluate (not evaluate_all).
        selected = await page.evaluate(
            """(args) => {
                const safeText = item => {
                    if (!item) return '';
                    return (item.innerText || item.getAttribute('aria-label') || item.value || '')
                        .replace(/\\s+/g, ' ').trim().toLocaleLowerCase();
                };
                const wantedName = safeText({ innerText: args.name });
                const wantedValue = safeText({ innerText: args.value });
                const nodes = [...document.querySelectorAll(
                    '.slicerContainer, [aria-label*="Slicer" i], [data-visual-type*="slicer" i], .visualContainer, [data-visual-container]'
                )];
                const container = nodes.find(node => safeText(node).includes(wantedName));
                if (!container) return false;
                const option = [...container.querySelectorAll('[role="option"], label, button, input')]
                    .find(node => safeText(node) === wantedValue);
                if (!option || option.disabled) return false;
                option.click();
                return true;
            }""",
            {"name": slicer_name, "value": value},
        )
        if selected:
            await self_wait_for_dashboard(page)
        return bool(selected)
    except Exception:
        logger.exception("Failed to apply slicer value | slicer=%s | value=%s", slicer_name, value)
        return False


async def self_wait_for_dashboard(page) -> None:
    """Wait briefly for a slicer action to repaint without reloading the page."""
    await page.wait_for_timeout(500)
    for _ in range(8):
        loading = await page.locator(".loading, [aria-label*='loading' i], [aria-busy='true']").count()
        if not loading:
            await page.wait_for_timeout(350)
            return
        await page.wait_for_timeout(350)
